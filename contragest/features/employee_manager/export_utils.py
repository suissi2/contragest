import os
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from contragest.core.excel_utils import safe_excel_cell_str
from fpdf import FPDF
from contragest.core.database import SessionLocal, AppConfig
from contragest.core.i18n import tr
from contragest.core.pdf_utils import safe_pdf_str

def _get_company_logo_path():
    session = SessionLocal()
    try:
        config = session.query(AppConfig).first()
        if config and config.company_logo_path and os.path.exists(config.company_logo_path):
            return config.company_logo_path
    except Exception:
        pass
    finally:
        session.close()
    return None

def export_to_excel(data, columns, filepath, title="Employee Report"):
    """
    Exports a list of dicts/rows to Excel.
    data: List of lists (row values)
    columns: List of strings (headers)
    """
    import logging
    logger = logging.getLogger(__name__)

    if not data:
        raise ValueError("No data to export.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Export"

    from openpyxl.styles import NamedStyle
    thin_border = Border(left=Side(style='thin', color="BFBFBF"), 
                         right=Side(style='thin', color="BFBFBF"), 
                         top=Side(style='thin', color="BFBFBF"), 
                         bottom=Side(style='thin', color="BFBFBF"))
                         
    data_style = NamedStyle(name="data_style")
    data_style.border = thin_border
    data_style.alignment = Alignment(horizontal='center', vertical='center')
    wb.add_named_style(data_style)

    # 1. Header Logic (Logo + Title)
    logo_path = _get_company_logo_path()
    start_row = 1
    if logo_path:
        try:
            from openpyxl.drawing.image import Image
            img = Image(logo_path)
            # Resize logo to fit roughly in a 60pt height row
            aspect_ratio = img.width / img.height
            img.height = 60
            img.width = 60 * aspect_ratio
            ws.add_image(img, 'A1')
            ws.row_dimensions[1].height = 65
            start_row = 2
        except:
            pass

    title_cell = ws.cell(row=start_row, column=1, value=title.upper())
    title_cell.font = Font(bold=True, size=14)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=max(len(columns), 1))
    title_cell.alignment = Alignment(horizontal='center')
    
    header_row = start_row + 2
    row_idx = header_row

    # 2. Table Headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    for col_idx, header in enumerate(columns, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=str(header).upper())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # 3. Table Data
    row_idx += 1
    for row in data:
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=safe_excel_cell_str(value) if value is not None else "-")
            cell.style = "data_style"
        row_idx += 1

    # Auto-adjust column widths
    for i, col in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    try:
        wb.save(filepath)
    except Exception as e:
        logger.exception("Failed to save Excel file")
        raise RuntimeError(f"Could not save Excel file. Is it already open? Error: {e}")

def export_to_pdf(data, columns, filepath, title="Employee Report"):
    """
    Exports a list of lists to PDF using FPDF.
    """
    if not data:
        raise ValueError("No data to export.")
        
    if not columns:
        # Avoid ZeroDivisionError if no columns are selected
        columns = ["Data"]
        data = [[str(row)] for row in data]

    class PDF(FPDF):
        def footer(self):
            self.set_y(-15)
            self.set_font("Helvetica", "I", 8)
            self.cell(0, 10, f"Page {self.page_no()}", align="C")

    # Orientation: Landscape if many columns
    orientation = 'L' if len(columns) > 7 else 'P'
    pdf = PDF(orientation=orientation, unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # 1. Header
    logo_path = _get_company_logo_path()
    if logo_path:
        try:
            pdf.image(logo_path, x=10, y=8, h=25)
        except:
            pass

    pdf.set_font("Helvetica", "B", 16)
    # Using new_x/new_y for fpdf2 compatibility
    pdf.cell(0, 15, safe_pdf_str(title).upper(), align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "I", 10)
    pdf.cell(0, 10, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(10)

    # 2. Table
    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(31, 78, 121)
    pdf.set_text_color(255, 255, 255)

    # Calculate column widths
    page_width = 190 if orientation == 'P' else 277 # Adjusted for margins
    col_width = page_width / len(columns)

    for h in columns:
        pdf.cell(col_width, 8, safe_pdf_str(h).upper(), border=1, align="C", fill=True)
    pdf.ln()

    pdf.set_font("Helvetica", "", 7)
    pdf.set_text_color(0, 0, 0)
    
    fill = False
    for row in data:
        # Check if we need a new page for this row
        if pdf.get_y() > (270 if orientation == 'P' else 180):
            pdf.add_page()
            # Redraw headers
            pdf.set_font("Helvetica", "B", 8)
            pdf.set_fill_color(31, 78, 121)
            pdf.set_text_color(255, 255, 255)
            for h in columns:
                pdf.cell(col_width, 8, safe_pdf_str(h).upper(), border=1, align="C", fill=True)
            pdf.ln()
            pdf.set_font("Helvetica", "", 7)
            pdf.set_text_color(0, 0, 0)

        pdf.set_fill_color(245, 245, 245) if fill else pdf.set_fill_color(255, 255, 255)
        for val in row:
            text = safe_pdf_str(val) if val is not None else "-"
            # Truncate if too long for the cell to avoid messy overlap
            max_chars = int(col_width * 1.5)
            if len(text) > max_chars:
                text = text[:max_chars-3] + "..."
            pdf.cell(col_width, 7, text, border=1, align="C", fill=True)
        pdf.ln()
        fill = not fill

    try:
        pdf.output(filepath)
    except Exception as e:
        # Re-raise with better context if output fails (e.g. file locked)
        raise RuntimeError(f"Could not save PDF file. Is it already open? Error: {e}")
