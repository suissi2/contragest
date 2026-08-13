import tkinter as tk
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from ttkbootstrap.dialogs import Messagebox, Querybox
from ttkbootstrap.widgets import DateEntry
from contragest.core.database import SessionLocal
from contragest.features.employee_manager.service import EmployeeService
from contragest.core.i18n import tr
from typing import Optional
import datetime
from PIL import Image, ImageTk
import os
from contragest.core.gui_utils import calculate_daily_password
from contragest.core.status_bar import StatusLabel
from contragest.features.pointage.service import PointageService
from tkinter import filedialog
import threading
from contragest.core.gui_utils import DesignTokens
import cv2
import numpy as np
from contragest.core.logging import setup_logger
from contragest.features.employee_manager import export_utils
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF

logger = setup_logger("employee_manager_ui")

# ─── COLOR TOKENS (NEBULA MIDNIGHT DESIGN SYSTEM) ────────────────────────────
PANEL_BG = DesignTokens.SURFACE
MAIN_BG = DesignTokens.BG_APP
ACCENT_BLUE = DesignTokens.PRIMARY
ACCENT_AMBER = DesignTokens.WARNING
TEXT_HIGH = DesignTokens.TEXT
TEXT_MUTED = DesignTokens.TEXT_MUTED
BORDER_COLOR = DesignTokens.SECONDARY
SUCCESS_EMERALD = DesignTokens.SUCCESS
DANGER_ROSE = DesignTokens.DANGER
# ─────────────────────────────────────────────────────────────────────────────

class WebcamCaptureDialog(ttk.Toplevel):
    def __init__(self, parent, emp_name, callback):
        super().__init__(parent)
        self.title(f"📸 CAPTURE IDENTITY - {emp_name}")
        self.geometry("660x600")
        self.resizable(False, False)
        self.center_window()
        
        self.callback = callback
        self.cap = None
        
        # Multiple Index Search & Backend Optimization (DSHOW is better on Windows)
        for idx in [0, 1, 2, 700]: # 700 is a common virtual camera offset
            try:
                temp_cap = cv2.VideoCapture(idx, cv2.CAP_DSHOW)
                if temp_cap.isOpened():
                    self.cap = temp_cap
                    logger.info(f"Webcam initialized on index {idx}")
                    break
            except:
                continue
        
        self._current_frame = None
        self._build_ui()
        
        if self.cap and self.cap.isOpened():
            self._update_feed()
        else:
            self.canvas_lbl.config(text="⚠️ NO CAMERA DETECTED\nPlease verify hardware connection\nor use 'UPLOAD PHOTO' instead.", 
                                 foreground=DANGER_ROSE, font=("Space Mono", 10, "bold"))
            self.cap_btn.config(state="disabled")
        
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def center_window(self):
        self.update_idletasks()
        w, h = 660, 600
        x = (self.winfo_screenwidth() // 2) - (w // 2)
        y = (self.winfo_screenheight() // 2) - (h // 2)
        self.geometry(f"{w}x{h}+{x}+{y}")

    def _build_ui(self):
        self.canvas_fr = ttk.Frame(self, bootstyle="dark", padding=2)
        self.canvas_fr.pack(fill=BOTH, expand=YES, padx=10, pady=10)
        
        self.canvas_lbl = tk.Label(self.canvas_fr, background="#000")
        self.canvas_lbl.pack(fill=BOTH, expand=YES)
        
        btn_fr = ttk.Frame(self)
        btn_fr.pack(fill=X, pady=15)
        
        self.cap_btn = ttk.Button(
            btn_fr, text="🚀 CAPTURE PHOTO", 
            bootstyle="success-glow", 
            command=self._capture,
            width=25
        )
        self.cap_btn.pack(side=LEFT, padx=120, expand=True)

    def _update_feed(self):
        if not self.cap or not self.cap.isOpened():
            return
            
        ret, frame = self.cap.read()
        if ret:
            # Mirror the frame for more natural experience
            frame = cv2.flip(frame, 1)
            self._current_frame = frame.copy()
            
            # Convert to RGB for PIL
            cv2_img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            img = Image.fromarray(cv2_img)
            img = img.resize((640, 480), Image.LANCZOS)
            
            self._tk_img = ImageTk.PhotoImage(image=img)
            self.canvas_lbl.config(image=self._tk_img)
            
        self._after_id = self.after(15, self._update_feed)

    def _capture(self):
        if self._current_frame is not None:
            # Final high-quality capture
            self.callback(self._current_frame)
            self._on_close()

    def _on_close(self):
        if hasattr(self, "_after_id"):
            self.after_cancel(self._after_id)
        if self.cap:
            self.cap.release()
        self.destroy()

class EmployeeManagerWindow(ttk.Toplevel):
    def __init__(self, parent, current_user=None):
        super().__init__(parent)
        self.current_user = current_user
        self.title("Employee Hierarchy Manager")
        self.geometry("1400x900")
        self.center_window()
        
        self.session = SessionLocal()
        self.service = EmployeeService(self.session)
        self.pointage_service = PointageService(self.session)
        self.selected_employee_id = None
        self.selected_employee_ids = []
        self.all_employees_data = [] # Cache for filtering
        self.page_size_var = ttk.StringVar(value="20")
        self.lbl_total_rows = None
        self.lbl_selected_rows = None
        
        # Field configuration (Display Name Key, DB Attribute, Default Visible)
        self.field_config = [
            ("SEL", "selection_icon", True),
            (tr("id"), "id", False),
            (tr("registration_number"), "registration_number", True),
            (tr("civility"), "civility", False),
            (tr("last_name"), "last_name", True),
            (tr("first_name"), "first_name", True),
            (tr("department"), "department_name", True),
            (tr("role_title_col") if tr("role_title_col") != "role_title_col" else "Role Title", "role_title", True),
            (tr("matrimonial_status"), "matrimonial_status", False),
            (tr("children_count"), "children_count", False),
            (tr("cnss"), "cnss", False),
            (tr("id_card_number"), "id_card_number", False),
            (tr("passport"), "passport", False),
            (tr("dob"), "dob", False),
            (tr("nationality"), "nationality", False),
            (tr("address"), "address", False),
            (tr("mobile_phone"), "mobile_phone", False),
            (tr("office_phone"), "office_phone", False),
            (tr("privilege"), "privilege", False),
            (tr("hire_date"), "hire_date", False),
            (tr("exit_date"), "exit_date", False),
            (tr("gross_salary"), "gross_salary", False),
            (tr("net_salary"), "net_salary", False),
            ("WEEKLY DAY OFF", "weekly_day_off", False),
            ("AUTO PUNCH", "is_auto_punch", False),
            ("PIC", "photo_icon", True)
        ]
        
        self.column_vars = {} # Stores BooleanVar for each checkbox
        self.active_attrs = []
        self.active_cols = []
        self._pointage_machines = {}   # display_name -> machine_id

        # Add Persistent Status Bar (Reserve bottom area before UI build)
        self.status_bar = StatusLabel(self)
        self.status_bar.pack(side=BOTTOM, fill=X)
        self.status_bar.set_status("Ready")

        self.create_widgets()
        self.load_all_employees()

    def center_window(self):
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f'{width}x{height}+{x}+{y}')

    def create_widgets(self):
        # Initialize hub variables
        self._pointage_machines_dict = {}
        self._dept_machine_var = ttk.StringVar()
        self._selected_dept_id = None

        # The main window now only contains the top-level Notebook for clean modularity
        self.right_notebook = ttk.Notebook(self)
        self.right_notebook.pack(fill=BOTH, expand=YES, padx=10, pady=10)

        # Tab 1: MASTER PERSONNEL HUB (THE ALL-IN-ONE COMMAND CENTER)
        self.grid_container = ttk.Frame(self.right_notebook)
        self.right_notebook.add(self.grid_container, text=" 👥  Personnel Command Center ")
        self._build_unified_personnel_hub()
        
        # Tab 2: Machine Sync & Personal (MOVED FROM POINTAGE)
        self.personal_sync_container = tk.Frame(self.right_notebook, background=MAIN_BG)
        self.right_notebook.add(self.personal_sync_container, text=" 👤  Machine Sync & Personal ")
        self._build_personal_sync_tab()

    def _build_unified_personnel_hub(self):
        """Combines Hierarchy, Toolbar, Filter Bar, Sidebar, and Grid into a single ultra-hub."""
        # 1. Internal Toolbar and Filter Bar (Top Area)
        hub_header = ttk.Frame(self.grid_container)
        hub_header.pack(fill=X, side=TOP)
        
        self.create_professional_toolbar(parent=hub_header)
        self.create_filter_bar(parent=hub_header)
        
        # 2. Master Paned Layout (The Three-Pane Center)
        self.hub_paned = ttk.Panedwindow(self.grid_container, orient=HORIZONTAL)
        self.hub_paned.pack(fill=BOTH, expand=YES, pady=(10, 0))
        
        # --- PANE 1: ORGANIZATIONAL HIERARCHY (LEFT) ---
        hierarchy_section = ttk.Frame(self.hub_paned)
        self.hub_paned.add(hierarchy_section, weight=2)
        
        hierarchy_fr = ttk.LabelFrame(hierarchy_section, text="Organizational Hierarchy")
        hierarchy_fr.pack(fill=BOTH, expand=YES, padx=5)
        
        # Machine Sync Mini-Panel
        m_sync_fr = ttk.Frame(hierarchy_fr, padding=(5, 2))
        m_sync_fr.pack(fill=X)
        self._dept_machine_combo = ttk.Combobox(m_sync_fr, textvariable=self._dept_machine_var, state="readonly", width=15)
        self._dept_machine_combo.pack(side=LEFT, padx=2)
        ttk.Button(m_sync_fr, text="🔄", width=3, bootstyle="link", command=self._refresh_dept_machine_combo).pack(side=LEFT)
        ttk.Button(m_sync_fr, text="🔃 Sync", bootstyle="info-link", command=self._import_depts_to_machine).pack(side=RIGHT)

        self.dept_tree = ttk.Treeview(hierarchy_fr, selectmode="browse", show="tree headings", bootstyle="dark")
        self.dept_tree["columns"] = ("employees")
        self.dept_tree.heading("#0", text="Department / Employee", anchor=W)
        self.dept_tree.heading("employees", text="Count", anchor=CENTER)
        self.dept_tree.column("#0", width=250, anchor=W)
        self.dept_tree.column("employees", width=60, anchor=CENTER)
        
        scroll_tree = ttk.Scrollbar(hierarchy_fr, orient=VERTICAL, command=self.dept_tree.yview)
        self.dept_tree.configure(yscrollcommand=scroll_tree.set)
        self.dept_tree.pack(side=TOP, fill=BOTH, expand=YES)
        scroll_tree.place(in_=self.dept_tree, relx=1.0, rely=0, relheight=1.0, anchor="ne")
        
        # Dept Action Buttons (Small & Clean)
        dept_btn_fr = ttk.Frame(hierarchy_fr, padding=5)
        dept_btn_fr.pack(fill=X, side=BOTTOM)
        ttk.Button(dept_btn_fr, text="🔄", width=3, bootstyle="info-link", command=self._global_refresh_and_reset).pack(side=LEFT, padx=2)
        ttk.Button(dept_btn_fr, text="➕", width=3, bootstyle="success-link", command=self._add_department).pack(side=LEFT, padx=2)
        ttk.Button(dept_btn_fr, text="✏️", width=3, bootstyle="warning-link", command=self._rename_department).pack(side=LEFT, padx=2)
        ttk.Button(dept_btn_fr, text="🗑️", width=3, bootstyle="danger-link", command=self._delete_department).pack(side=LEFT, padx=2)
        
        # Initialize Hierarchy Data
        self._refresh_dept_tree()
        self._refresh_dept_machine_combo()

        self.dept_tree.bind("<<TreeviewSelect>>", self._on_dept_select_hub)
        self.dept_tree.bind("<Button-3>", self._on_dept_right_click)
        self.dept_tree.bind("<Double-1>", self._on_dept_double_click)

        # --- PANE 2: PERSONNEL IDENTITY (MIDDLE) ---
        sidebar_container = ttk.Frame(self.hub_paned)
        self.hub_paned.add(sidebar_container, weight=1)
        
        # Identity Preview Pane
        self.photo_preview_fr = ttk.LabelFrame(sidebar_container, text="Identity Preview")
        self.photo_preview_fr.pack(fill=X, padx=5, pady=(0, 10))
        
        self.photo_label = ttk.Label(self.photo_preview_fr, text="[No Selection]", anchor=CENTER, padding=10)
        self.photo_label.pack(fill=X, expand=YES)
        
        action_fr = ttk.Frame(self.photo_preview_fr, padding=5)
        action_fr.pack(fill=X)
        ttk.Button(action_fr, text="📸", width=3, bootstyle="info-link", command=self._capture_photo_for_selected).pack(side=LEFT, padx=5)
        ttk.Button(action_fr, text="📁", width=3, bootstyle="info-link", command=self._upload_photo_for_selected).pack(side=LEFT, padx=5)
        ttk.Button(action_fr, text="🗑️", width=3, bootstyle="danger-link", command=self._remove_photo_for_selected).pack(side=LEFT, padx=5)

        # Field Visibility Pane
        selection_frame = ttk.LabelFrame(sidebar_container, text="Data Visibility")
        selection_frame.pack(fill=BOTH, expand=YES, padx=5)
        
        btn_batch_frame = ttk.Frame(selection_frame)
        btn_batch_frame.pack(fill=X, pady=(0, 5))
        ttk.Button(btn_batch_frame, text="All", bootstyle="link", command=self.select_all_fields).pack(side=LEFT)
        ttk.Button(btn_batch_frame, text="None", bootstyle="link", command=self.deselect_all_fields).pack(side=LEFT, padx=5)
        
        scroll_container = ttk.Frame(selection_frame)
        scroll_container.pack(fill=BOTH, expand=YES)
        canvas = ttk.Canvas(scroll_container, highlightthickness=0)
        scrollbar = ttk.Scrollbar(scroll_container, orient=VERTICAL, command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=LEFT, fill=BOTH, expand=YES)
        scrollbar.pack(side=RIGHT, fill=Y)

        for label, attr, default_val in self.field_config:
            var = ttk.BooleanVar(value=default_val)
            self.column_vars[attr] = var
            ttk.Checkbutton(scrollable_frame, text=label, variable=var, bootstyle="round-toggle", command=self.update_table_columns).pack(anchor=W, pady=2, padx=5)

        # --- PANE 3: ACTIVE RECORDS (RIGHT) ---
        grid_section = ttk.Frame(self.hub_paned)
        self.hub_paned.add(grid_section, weight=5)
        
        self.grid_frame = ttk.LabelFrame(grid_section, text="Personnel Database")
        self.grid_frame.pack(fill=BOTH, expand=YES, padx=5)

        # Identity Quick Actions Toolbar
        identity_toolbar = ttk.Frame(self.grid_frame)
        identity_toolbar.pack(fill=X, padx=10, pady=(5, 0))
        ttk.Label(identity_toolbar, text="Selection Identity:", font=("Inter", 9, "bold")).pack(side=LEFT)
        ttk.Button(identity_toolbar, text="📸 Capture", command=self._capture_photo_for_selected, bootstyle="info-link").pack(side=LEFT, padx=8)
        ttk.Button(identity_toolbar, text="📁 Upload", command=self._upload_photo_for_selected, bootstyle="info-link").pack(side=LEFT, padx=8)
        ttk.Button(identity_toolbar, text="🗑️ Remove", command=self._remove_photo_for_selected, bootstyle="danger-link").pack(side=LEFT, padx=8)
        
        self.table_container = ttk.Frame(self.grid_frame)
        self.table_container.pack(fill=BOTH, expand=YES)
        
        self.update_table_columns()
        self.create_table_footer(self.grid_frame)

    def _on_dept_select_hub(self, event):
        """Unified selection handler that filters the table based on hierarchy selection."""
        selection = self.dept_tree.selection()
        if not selection:
            return
            
        item_iid = selection[0]
        tags = self.dept_tree.item(item_iid, "tags")
        text = self.dept_tree.item(item_iid, "text").replace("  📁  ", "").replace("  👤  ", "").strip()
        
        if tags and "emp_node" in tags:
            # It's an employee node
            emp_id_str = tags[0].replace("emp_", "")
            self.on_table_select(None, manual_emp_id=int(emp_id_str))
        else:
            # It's a department node
            self._selected_dept_id = int(tags[0]) if tags else None
            self.dept_filter_var.set(text)
            self.search_employees()
            # Still update photo preview to clear it or show nothing
            self._update_photo_preview(None)

    def create_professional_toolbar(self, parent=None):
        # 1. Header with Logo & Title (Only if parent is global/none)
        if not parent:
            header_container = ttk.Frame(self, bootstyle="light")
            header_container.pack(fill=X, side=TOP)
            
            self.logo_lbl = ttk.Label(header_container, bootstyle="inverse-light")
            self.logo_lbl.pack(side=LEFT, padx=15, pady=5)
            self._load_logo()
            
            title_f = ttk.Frame(header_container, bootstyle="light")
            title_f.pack(side=LEFT)
            ttk.Label(title_f, text="EMPLOYEE MANAGEMENT HUB", font=("Inter", 14, "bold"), bootstyle="inverse-light").pack(anchor=W)
            ttk.Label(title_f, text="Advanced Hierarchy & Machine Synchronization", font=("Inter", 9), bootstyle="inverse-light").pack(anchor=W)

        # 2. Toolbar Actions
        toolbar_container = ttk.Frame(parent if parent else self, bootstyle="light" if not parent else "default")
        toolbar_container.pack(fill=X, side=TOP)
        
        # --- Employees Pane ---
        emp_pane = ttk.LabelFrame(toolbar_container, text="Employees")
        emp_pane.pack(side=LEFT, padx=5)
        
        emp_actions = [
             ("➕", tr("add"), self.add_emp, SUCCESS),
             ("📥", tr("import"), self.import_data, INFO),
             ("🏢", tr("department_manager"), self.manage_groups, INFO),
             ("🔄", "Sync/Personal", self.show_personal_sync_tab, INFO),
             ("☑", "Select All", self.select_all_employees, INFO),
             ("☐", "Deselect All", self.deselect_all_employees, SECONDARY)
        ]
        for icon, label, cmd, bstyle in emp_actions:
            ttk.Button(emp_pane, text=f"{icon} {label}", command=cmd, bootstyle=bstyle).pack(side=LEFT, padx=4, pady=4)

        # --- Archive Pane ---
        arch_pane = ttk.LabelFrame(toolbar_container, text="Archive")
        arch_pane.pack(side=LEFT, padx=5)
        
        arch_actions = [
            ("🗂️", "Archive", self.archive_selected_employee, WARNING),
            ("📂", "View Archive", self.open_archive_panel, INFO)
        ]
        for icon, label, cmd, bstyle in arch_actions:
            ttk.Button(arch_pane, text=f"{icon} {label}", command=cmd, bootstyle=bstyle).pack(side=LEFT, padx=4, pady=4)

        # --- Identity Pane ---
        id_pane = ttk.LabelFrame(toolbar_container, text="Identity")
        id_pane.pack(side=LEFT, padx=5)
        
        id_actions = [
            ("📸", "Capture", self._capture_photo_for_selected, INFO),
            ("📁", "Upload", self._upload_photo_for_selected, INFO),
            ("🗑️", "Remove", self._remove_photo_for_selected, DANGER)
        ]
        for icon, label, cmd, bstyle in id_actions:
            ttk.Button(id_pane, text=f"{icon} {label}", command=cmd, bootstyle=bstyle).pack(side=LEFT, padx=4, pady=4)

        # --- Export Pane ---
        exp_pane = ttk.LabelFrame(toolbar_container, text="Export")
        exp_pane.pack(side=LEFT, padx=5)

        exp_actions = [
            ("📊", "Excel", self.export_selected_excel, SUCCESS),
            ("📕", "PDF", self.export_selected_pdf, DANGER)
        ]
        for icon, label, cmd, bstyle in exp_actions:
            ttk.Button(exp_pane, text=f"{icon} {label}", command=cmd, bootstyle=bstyle).pack(side=LEFT, padx=4, pady=4)

    def _load_logo(self):
        """Dynamic logo loading for Employee Manager."""
        from contragest.core.database import SessionLocal, AppConfig
        session = SessionLocal()
        try:
            config = session.query(AppConfig).first()
            if config and config.company_logo_path and os.path.exists(config.company_logo_path):
                img = Image.open(config.company_logo_path)
                img.thumbnail((36, 36))
                self.logo_img = ImageTk.PhotoImage(img)
                self.logo_lbl.config(image=self.logo_img)
        except: pass
        finally: session.close()

    def create_table_footer(self, parent):
        footer = ttk.Frame(parent, bootstyle="secondary", padding=(10, 5))
        footer.pack(fill=X, side=BOTTOM)

        # Pagination / Rows per page
        ttk.Label(footer, text="Rows per page:", bootstyle="inverse-secondary").pack(side=LEFT, padx=(0, 5))
        
        self.page_size_combo = ttk.Combobox(
            footer, 
            textvariable=self.page_size_var, 
            values=["20", "50", "100", "150", "200", "250", "300"],
            width=5,
            state="readonly"
        )
        self.page_size_combo.pack(side=LEFT)
        self.page_size_combo.bind("<<ComboboxSelected>>", lambda e: self.update_table_columns())

        ttk.Separator(footer, orient=VERTICAL).pack(side=LEFT, fill=Y, padx=15)

        # Stats
        self.lbl_total_rows = ttk.Label(footer, text="Total: 0", bootstyle="inverse-secondary")
        self.lbl_total_rows.pack(side=LEFT, padx=5)

        ttk.Separator(footer, orient=VERTICAL).pack(side=LEFT, fill=Y, padx=15)

        self.lbl_selected_rows = ttk.Label(footer, text="Selected: 0", bootstyle="inverse-secondary")
        self.lbl_selected_rows.pack(side=LEFT, padx=5)

    def create_filter_bar(self, parent=None):
        filter_container = ttk.Frame(parent if parent else self, padding=10)
        filter_container.pack(fill=X)
        
        # Keyword Search
        ttk.Label(filter_container, text="🔍 Keyword:").pack(side=LEFT, padx=5, pady=5)
        self.search_var = ttk.StringVar()
        self.search_entry = ttk.Entry(filter_container, textvariable=self.search_var, width=25)
        self.search_entry.pack(side=LEFT, padx=5, pady=5)
        self.search_entry.bind("<Return>", lambda e: self.search_employees())
        
        # Department Filter
        ttk.Label(filter_container, text="🏢 Department:", bootstyle="inverse-secondary").pack(side=LEFT, padx=(15, 5), pady=5)
        self.dept_filter_var = ttk.StringVar()
        self._cb_dept = ttk.Combobox(filter_container, textvariable=self.dept_filter_var, width=20)
        self._cb_dept.pack(side=LEFT, padx=5, pady=5)
        
        # Role Title Filter
        ttk.Label(filter_container, text="🏷️ Role Title:", bootstyle="inverse-secondary").pack(side=LEFT, padx=(15, 5), pady=5)
        self.role_filter_var = ttk.StringVar()
        self._cb_role = ttk.Combobox(filter_container, textvariable=self.role_filter_var, width=20)
        self._cb_role.pack(side=LEFT, padx=5, pady=5)
        
        # Action Buttons
        ttk.Button(filter_container, text="✅ Apply Filters", bootstyle="primary", command=self.search_employees).pack(side=LEFT, padx=(20, 5), pady=5)
        ttk.Button(filter_container, text="🔄 Clear", bootstyle="light", command=self.clear_filters).pack(side=LEFT, padx=5, pady=5)

    def _populate_filter_dropdowns(self):
        """Extract unique values from cached employee data for comboboxes with autocomplete."""
        dept_list = sorted(list(set(e.get("department_name", "") for e in self.all_employees_data if e.get("department_name") and e.get("department_name") != "Unassigned")))
        role_list = sorted(list(set(e.get("role_title", "") for e in self.all_employees_data if e.get("role_title") and e.get("role_title") != "-")))
        
        self._filter_full_lists = {
            "departments": dept_list,
            "roles": role_list
        }
        
        if hasattr(self, "_cb_dept"):
            self._cb_dept.configure(values=self._filter_full_lists["departments"])
            self._cb_dept.bind("<KeyRelease>", lambda e: self._on_combobox_key(e, self._cb_dept, self._filter_full_lists["departments"]))
            self._cb_dept.bind("<ButtonPress-1>", lambda e: self._on_combobox_click(self._cb_dept, self._filter_full_lists["departments"]))

        if hasattr(self, "_cb_role"):
            self._cb_role.configure(values=self._filter_full_lists["roles"])
            self._cb_role.bind("<KeyRelease>", lambda e: self._on_combobox_key(e, self._cb_role, self._filter_full_lists["roles"]))
            self._cb_role.bind("<ButtonPress-1>", lambda e: self._on_combobox_click(self._cb_role, self._filter_full_lists["roles"]))

    def _on_combobox_key(self, event, combobox, full_list):
        if event.keysym in ('Up', 'Down', 'Return', 'Left', 'Right', 'Tab', 'Shift_L', 'Shift_R', 'Control_L', 'Control_R', 'Alt_L', 'Alt_R'):
            return
            
        typed_text = combobox.get().lower()
        if not typed_text:
            combobox.configure(values=full_list)
        else:
            filtered = [item for item in full_list if typed_text in str(item).lower()]
            combobox.configure(values=filtered)
            
        combobox.after("idle", lambda: combobox.event_generate('<Down>'))
        
    def _on_combobox_click(self, combobox, full_list):
        if not combobox.get():
            combobox.configure(values=full_list)

    def select_all_fields(self):
        for var in self.column_vars.values():
            var.set(True)
        self.update_table_columns()

    def deselect_all_fields(self):
        for var in self.column_vars.values():
            var.set(False)
        self.update_table_columns()

    # ── Machine import helpers ─────────────────────────────────────────────

    def _load_machine_combo(self):
        """Refresh the machine combobox from the pointage service."""
        try:
            from contragest.features.pointage.service import PointageService
            svc = PointageService(self.session)
            machines = svc.get_all_machines()
            self._pointage_machines = {
                f"{m.name} ({m.ip_address})": m.id for m in machines
            }
            self._emp_machine_combo["values"] = list(self._pointage_machines.keys())
            if self._pointage_machines:
                self._emp_machine_combo.current(0)
        except Exception:
            self._pointage_machines = {}

    def _get_selected_machine_id(self):
        key = self._emp_machine_var.get()
        return self._pointage_machines.get(key)

    def _import_employees_to_machine(self):
        """Push all employees to the selected pointage machine."""
        mid = self._get_selected_machine_id()
        if not mid:
            Messagebox.show_warning(tr("select_machine_first"), tr("information"))
            return
        try:
            from contragest.features.pointage.service import PointageService
            svc = PointageService(self.session)
            success, failed = svc.push_all_employees_to_machine(mid)
            msg = tr("import_machine_success").replace("{count}", str(success))
            if failed:
                msg += f"\n({failed} {tr('error').lower()}s)"
            Messagebox.show_info(msg, tr("success"))
        except Exception as e:
            Messagebox.show_error(
                tr("import_machine_failed").replace("{error}", str(e)),
                tr("error")
            )


    def transfer_employee(self):
        """Intelligent transfer action based on current selection."""
        # Check Departments Tab first
        if self.right_notebook.index("current") == 0:
            selection = self.dept_tree.selection()
            if not selection:
                Messagebox.show_warning("Please select an employee to transfer.", "No Selection")
                return
            tags = self.dept_tree.item(selection[0], "tags")
            if "emp_node" in tags:
                emp_id = int(tags[0].replace("emp_", ""))
                self._transfer_specific_employee(emp_id)
            else:
                Messagebox.show_warning("Please select an employee node, not a department.", "Invalid Selection")
            return

        # Check Employee Records Tab
        if not self.selected_employee_id:
            Messagebox.show_warning("Please select an employee to transfer.", "No Selection")
            return
        self._transfer_specific_employee(self.selected_employee_id)

    def _transfer_specific_employee(self, emp_id):
        """Opens a dialog to select the target department for transfer."""
        from contragest.core.database import Department, Employee
        emp = self.session.query(Employee).get(emp_id)
        if not emp: return

        # Create a simple selection dialog
        dialog = ttk.Toplevel(self)
        dialog.title(f"Transfer: {emp.last_name} {emp.first_name}")
        dialog.geometry("400x200")
        dialog.resizable(False, False)
        dialog.grab_set()
        
        main_fr = ttk.Frame(dialog, padding=20)
        main_fr.pack(fill=BOTH, expand=YES)
        
        ttk.Label(main_fr, text="Target Department:", font=("Inter", 10, "bold")).pack(anchor=W, pady=(0, 10))
        
        depts = self.session.query(Department).all()
        dept_map = {d.name: d.id for d in depts}
        
        dept_var = ttk.StringVar()
        combo = ttk.Combobox(main_fr, textvariable=dept_var, values=list(dept_map.keys()), state="readonly", width=40)
        combo.pack(fill=X, pady=(0, 20))
        
        # Pre-select current dept if exists
        if emp.dept_obj:
            combo.set(emp.dept_obj.name)
        elif dept_map:
            combo.current(0)

        def confirm_transfer():
            target_name = dept_var.get()
            target_id = dept_map.get(target_name)
            if target_id:
                if self._perform_transfer(emp_id, target_id):
                    dialog.destroy()

        btn_fr = ttk.Frame(main_fr)
        btn_fr.pack(fill=X)
        ttk.Button(btn_fr, text="CONFIRM TRANSFER", bootstyle="success", command=confirm_transfer).pack(side=RIGHT)
        ttk.Button(btn_fr, text="CANCEL", bootstyle="secondary-outline", command=dialog.destroy).pack(side=RIGHT, padx=10)

        # Center Dialog
        dialog.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() // 2) - (400 // 2)
        y = self.winfo_y() + (self.winfo_height() // 2) - (200 // 2)
        dialog.geometry(f"+{x}+{y}")

    def _perform_transfer(self, emp_id, target_id):
        """Backend execution of employee transfer."""
        try:
            self.service.update_employee(emp_id, department_id=target_id)
            self._refresh_dept_tree()
            self.refresh_table_data()
            self.status_bar.set_status(f"Employee {emp_id} successfully transferred.")
            return True
        except Exception as e:
            Messagebox.show_error(f"Transfer failed: {e}", "Error")
            return False

    def _on_dept_drag_start(self, event):
        item = self.dept_tree.identify_row(event.y)
        if not item: return
        tags = self.dept_tree.item(item, "tags")
        if tags and "emp_node" in tags:
            self._drag_data = {"item": item, "emp_id": int(tags[0].replace("emp_", ""))}
        else:
            self._drag_data = None

    def _on_dept_drag_stop(self, event):
        if not hasattr(self, "_drag_data") or not self._drag_data: return
        
        target_item = self.dept_tree.identify_row(event.y)
        if not target_item: return
        
        target_tags = self.dept_tree.item(target_item, "tags")
        target_dept_id = None
        
        if target_tags and "dept_node" in target_tags:
            target_dept_id = int(target_tags[0])
        elif target_tags and "emp_node" in target_tags:
            parent = self.dept_tree.parent(target_item)
            if parent:
                parent_tags = self.dept_tree.item(parent, "tags")
                if parent_tags and "dept_node" in parent_tags:
                    target_dept_id = int(parent_tags[0])
        
        if target_dept_id:
            from contragest.core.database import Employee
            emp = self.session.query(Employee).get(self._drag_data["emp_id"])
            if emp and emp.department_id != target_dept_id:
                emp_name = f"{emp.last_name} {emp.first_name}"
                dept_name = self.dept_tree.item(target_item, "text").strip().replace("📁", "").replace("👤", "").strip()
                if Messagebox.yesno(f"Transfer {emp_name} to {dept_name}?", "Confirm Transfer", parent=self):
                    self._perform_transfer(self._drag_data["emp_id"], target_dept_id)
        
        self._drag_data = None

    def _on_dept_right_click(self, event):
        item = self.dept_tree.identify_row(event.y)
        if not item: return
        self.dept_tree.selection_set(item)
        tags = self.dept_tree.item(item, "tags")
        
        menu = tk.Menu(self, tearoff=0)
        if tags and "emp_node" in tags:
            emp_id = int(tags[0].replace("emp_", ""))
            menu.add_command(label="✏️ Edit Profile", command=self.save_employee)
            menu.add_command(label="📦 Transfer Employee", command=lambda: self._transfer_specific_employee(emp_id))
            
            # Photo Submenu
            photo_menu = tk.Menu(menu, tearoff=0)
            photo_menu.add_command(label="📸 Capture from Webcam", command=self._capture_photo_for_selected)
            photo_menu.add_command(label="📁 Upload from File", command=self._upload_photo_for_selected)
            photo_menu.add_command(label="🗑️ Remove Photo", command=self._remove_photo_for_selected)
            menu.add_cascade(label="🖼️ Identity Photo", menu=photo_menu)
            
            menu.add_separator()
            menu.add_command(label="📤 Export to Machine", command=self._export_selected_employee)
        elif tags and "dept_node" in tags:
            menu.add_command(label="➕ Add Sub-Department", command=self._add_department)
            menu.add_command(label="✏️ Rename Department", command=self._rename_department)
            menu.add_command(label="🗑️ Delete Department", command=self._delete_department)
            
        menu.post(event.x_root, event.y_root)

    def _on_grid_right_click(self, event):
        """Right-click menu for the main employee grid."""
        row = self.emp_table.view.identify_row(event.y)
        if not row: return
        
        # Select the row if not already selected
        if row not in self.emp_table.view.selection():
            self.emp_table.view.selection_set(row)
            self.on_table_select(None)
            
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="✏️ Edit Employee", command=self.save_employee)
        menu.add_command(label="📦 Transfer Employee", command=self.transfer_employee)
        
        # Photo Submenu
        photo_menu = tk.Menu(menu, tearoff=0)
        photo_menu.add_command(label="📸 Capture from Webcam", command=self._capture_photo_for_selected)
        photo_menu.add_command(label="📁 Upload from File", command=self._upload_photo_for_selected)
        photo_menu.add_command(label="🗑️ Remove Photo", command=self._remove_photo_for_selected)
        menu.add_cascade(label="🖼️ Identity Photo", menu=photo_menu)
        
        menu.add_separator()
        menu.add_command(label="📤 Export to Machine", command=self._export_selected_employee)
        menu.add_command(label="🛑 Terminate Contract", command=self.terminate_employee)
        menu.add_command(label="📁 Archive Employee", command=self.archive_selected_employee)
        
        menu.post(event.x_root, event.y_root)

    def _remove_photo_for_selected(self):
        """Remove identity photo for selected employee."""
        if not self.selected_employee_id:
            Messagebox.show_warning("Please select an employee.", "No Selection")
            return
            
        if Messagebox.yesno("Are you sure you want to remove this employee's photo?", "Confirm Removal"):
            try:
                self.service.update_employee(self.selected_employee_id, photo_path=None)
                self._update_photo_preview(self.selected_employee_id)
                self.load_all_employees() # Refresh grid icon
                self.status_bar.set_status("Photo removed successfully.")
            except Exception as e:
                Messagebox.show_error(f"Failed to remove photo: {e}", "Error")

    def _update_photo_preview(self, emp_id=None):
        """Update the identity photo in the left sidebar."""
        if not emp_id:
            self.photo_label.configure(image="", text="[No Selection]")
            return
            
        emp = self.service.get_employee_full(emp_id)
        if not emp or not emp.photo_path or not os.path.exists(emp.photo_path):
            self.photo_label.configure(image="", text="[No Photo]")
            return
            
        try:
            from PIL import Image, ImageTk
            img = Image.open(emp.photo_path)
            # Standard headshot ratio, fits sidebar width
            img.thumbnail((160, 200))
            self._current_sidebar_img = ImageTk.PhotoImage(img)
            self.photo_label.configure(image=self._current_sidebar_img, text="")
        except Exception:
            self.photo_label.configure(image="", text="[Load Error]")

    def _on_dept_double_click(self, event):
        """Handle double-click on an employee node to open edit form."""
        item = self.dept_tree.identify_row(event.y)
        if not item: return
        tags = self.dept_tree.item(item, "tags")
        if tags and "emp_node" in tags:
            self.selected_employee_id = int(tags[0].replace("emp_", ""))
            self.save_employee()

    def _export_selected_employee(self):
        """Export the currently selected employee to all active terminals."""
        if not self.selected_employee_id:
            Messagebox.show_warning("Please select an employee first.", tr("information"))
            return

        # Check that employee has a registration number
        emp = self.service.get_employee_full(self.selected_employee_id)
        if not emp or not emp.registration_number:
            Messagebox.show_warning(tr("no_reg_number"), tr("information"))
            return

        Messagebox.show_info(tr("export_in_progress"), tr("information"))

        from contragest.features.pointage.sync_bus import sync_bus
        sync_bus.publish_employee_export(
            employee_id=self.selected_employee_id,
            callback=self._export_employee_callback,
            tk_root=self
        )

    def _export_employee_callback(self, success, message):
        """Callback invoked on main thread after export completes."""
        if success:
            parts = message.split("|")
            name = parts[1] if len(parts) > 1 else ""
            count = parts[2] if len(parts) > 2 else "0"
            msg = tr("export_success").replace("{NAME}", name).replace("{COUNT}", count)
            Messagebox.show_info(msg, tr("success"))
        else:
            parts = message.split("|")
            name = parts[1] if len(parts) > 1 else ""
            error = parts[2] if len(parts) > 2 else message
            msg = tr("export_failed").replace("{NAME}", name).replace("{ERROR}", error)
            Messagebox.show_error(msg, tr("error"))

    def terminate_employee(self):
        if not self.selected_employee_id:
            Messagebox.show_warning("Please select an employee first.", "No Selection")
            return
        
        import datetime
        today_str = datetime.date.today().strftime("%Y-%m-%d")
        
        exit_date_str = Querybox.get_string(
            prompt="Enter exit date (YYYY-MM-DD):",
            title="Terminate Employee",
            initial_value=today_str,
            parent=self
        )
        if not exit_date_str:
            return  # Cancelled
            
        try:
            exit_date = datetime.datetime.strptime(exit_date_str.strip(), "%Y-%m-%d").date()
        except ValueError:
            Messagebox.show_error("Invalid date format. Please use YYYY-MM-DD.", "Error", parent=self)
            return
            
        try:
            self.service.update_employee(self.selected_employee_id, exit_date=exit_date)
            self._refresh_dept_tree()
            self.refresh_table_data()
            self.status_bar.set_status(f"Employee {self.selected_employee_id} terminated on {exit_date_str}.")
            Messagebox.show_info(f"Employee terminated on {exit_date_str} successfully.", "Success", parent=self)
        except Exception as e:
            Messagebox.show_error(f"Failed to terminate employee: {e}", "Error", parent=self)

    def cancel_edit(self):
        self.selected_employee_id = None
        if hasattr(self, 'emp_table'):
            self.emp_table.view.selection_remove(self.emp_table.view.selection())
        self.clear_filters()

    def _global_refresh_and_reset(self):
        """Unified reset: clear search, reload DB, and refresh hierarchy."""
        self.clear_filters()
        self._refresh_dept_tree()
        self._refresh_dept_machine_combo()
        self._update_photo_preview(None)
        self._current_selected_emp_id = None
        if hasattr(self, "dept_tree"):
            self.dept_tree.selection_remove(self.dept_tree.selection())

    def clear_filters(self):
        """Reset search filters and reload all active employees."""
        self.search_var.set("")
        if hasattr(self, "dept_filter_var"):
            self.dept_filter_var.set("")
        if hasattr(self, "role_filter_var"):
            self.role_filter_var.set("")
        self._populate_filter_dropdowns()
        self.refresh_table_data()

    def manage_groups(self):
        """Switch to the Unified Personnel Hub."""
        self.right_notebook.select(self.grid_container)

    def _build_departments_tab(self):
        """Integrated Department Manager UI logic."""
        parent = self.dept_container
        self._selected_dept_id = None
        self._pointage_machines_dict = {}

        # 1. Header (OLED Styling)
        header = ttk.Frame(parent, bootstyle="info")
        header.pack(fill=X)
        ttk.Label(
            header, text=f"  🏢  {tr('department_manager').upper()}",
            font=("Inter", 12, "bold"),
            bootstyle="inverse-info", padding=(15, 10)
        ).pack(side=LEFT)

        # 2. Machine Import Panel
        machine_frame = ttk.LabelFrame(parent, text=f"🖥️ {tr('import_to_machine')}")
        machine_frame.pack(fill=X, padx=15, pady=(15, 0))

        ttk.Label(machine_frame, text=f"{tr('machine_name')}:", font=("Inter", 9)).pack(side=LEFT, padx=(10, 6), pady=10)

        self._dept_machine_var = ttk.StringVar()
        self._dept_machine_combo = ttk.Combobox(
            machine_frame, textvariable=self._dept_machine_var,
            width=35, state="readonly"
        )
        self._dept_machine_combo.pack(side=LEFT, padx=(0, 10), pady=10)

        ttk.Button(
            machine_frame, text="🔄", bootstyle="outline-info", width=3,
            command=self._refresh_dept_machine_combo
        ).pack(side=LEFT, padx=(0, 10), pady=10)

        ttk.Button(
            machine_frame,
            text=f"🔃 {tr('import_depts_to_machine')}",
            bootstyle="success-glow",
            command=self._import_depts_to_machine
        ).pack(side=LEFT, pady=10)

        # 3. Main Content (Treeview)
        tree_container = ttk.Frame(parent, padding=15)
        tree_container.pack(fill=BOTH, expand=YES)

        self.dept_tree = ttk.Treeview(tree_container, selectmode="browse", show="tree headings", bootstyle="dark")
        self.dept_tree["columns"] = ("reg", "role", "employees")
        self.dept_tree.heading("#0", text=tr("department") + " / " + tr("employee"), anchor=W)
        self.dept_tree.heading("reg", text=tr("registration_number"), anchor=W)
        self.dept_tree.heading("role", text=tr("role_title_col") if tr("role_title_col") != "role_title_col" else "Role Title", anchor=W)
        self.dept_tree.heading("employees", text=tr("employees_count"), anchor=CENTER)
        
        self.dept_tree.column("#0", width=350, anchor=W)
        self.dept_tree.column("reg", width=120, anchor=W)
        self.dept_tree.column("role", width=250, anchor=W)
        self.dept_tree.column("employees", width=100, anchor=CENTER)

        scrollbar = ttk.Scrollbar(tree_container, orient=VERTICAL, command=self.dept_tree.yview)
        self.dept_tree.configure(yscrollcommand=scrollbar.set)
        self.dept_tree.pack(side=LEFT, fill=BOTH, expand=YES)
        scrollbar.pack(side=RIGHT, fill=Y)

        self.dept_tree.bind("<<TreeviewSelect>>", self._on_dept_select)
        self.dept_tree.bind("<ButtonPress-1>", self._on_dept_drag_start)
        self.dept_tree.bind("<ButtonRelease-1>", self._on_dept_drag_stop)
        self.dept_tree.bind("<Button-3>", self._on_dept_right_click)
        self.dept_tree.bind("<Double-1>", self._on_dept_double_click)

        # 4. Action Buttons
        btn_frame = ttk.Frame(parent, padding=15)
        btn_frame.pack(fill=X, side=BOTTOM)

        ttk.Button(
            btn_frame, text=f"➕ {tr('add_department')}",
            bootstyle="success", command=self._add_department, width=20
        ).pack(side=LEFT, padx=5)

        ttk.Button(
            btn_frame, text=f"✏️ {tr('rename_department')}",
            bootstyle="warning", command=self._rename_department, width=20
        ).pack(side=LEFT, padx=5)

        ttk.Button(
            btn_frame, text=f"📦 {tr('transfer').upper()}",
            bootstyle="info-outline", command=self.transfer_employee, width=20
        ).pack(side=LEFT, padx=5)

        ttk.Button(
            btn_frame, text=f"🗑️ {tr('delete_department')}",
            bootstyle="danger", command=self._delete_department, width=20
        ).pack(side=LEFT, padx=5)

        # Initial Load
        self._refresh_dept_tree()
        self._refresh_dept_machine_combo()

    def _refresh_dept_machine_combo(self):
        try:
            machines = self.pointage_service.get_all_machines()
            self._pointage_machines_dict = {
                f"{m.name} ({m.ip_address})": m.id for m in machines
            }
            self._dept_machine_combo["values"] = list(self._pointage_machines_dict.keys())
            if self._pointage_machines_dict:
                self._dept_machine_combo.current(0)
        except: self._pointage_machines_dict = {}

    def _import_depts_to_machine(self):
        mid = self._pointage_machines_dict.get(self._dept_machine_var.get())
        if not mid:
            Messagebox.show_warning(tr("select_machine_first"), tr("information"), parent=self)
            return
        try:
            synced, failed = self.pointage_service.push_all_departments_to_machine(mid)
            msg = tr("dept_sync_saved").replace("{count}", str(synced))
            if failed: msg += f"\n({failed} errors)"
            Messagebox.show_info(msg, tr("success"), parent=self)
        except Exception as e:
            Messagebox.show_error(f"Sync failed: {e}", tr("error"), parent=self)

    def _refresh_dept_tree(self):
        for item in self.dept_tree.get_children():
            self.dept_tree.delete(item)
        
        # 1. Sort top-level departments alphabetically
        raw_depts = self.service.get_department_hierarchy()
        top_depts = sorted(raw_depts, key=lambda d: str(d.name or "").lower())
        
        for dept in top_depts:
            self._insert_dept_node("", dept)

    def _insert_dept_node(self, parent_iid, dept):
        """Recursively insert departments and their nested active employees into the treeview."""
        # Filter out archived employees
        active_emps = [e for e in (dept.employees if dept.employees else []) if not e.is_archived]
        emp_count = len(active_emps)
        
        # Insert Department Node
        iid = self.dept_tree.insert(
            parent_iid, "end",
            text=f"  📁  {dept.name}",
            values=(emp_count,),
            tags=(str(dept.id), "dept_node")
        )
        
        # 2. Insert Employee Nodes (sorted alphabetically by Last Name -> First Name)
        sorted_emps = sorted(active_emps, key=lambda e: (
            str(e.last_name or "").strip().lower(),
            str(e.first_name or "").strip().lower()
        ))
        for emp in sorted_emps:
            reg = emp.registration_number or "-"
            full_name = f"{emp.last_name} {emp.first_name}" if emp.last_name else (emp.first_name or "Unknown")
            role = emp.role_title or "-"
            
            self.dept_tree.insert(
                iid, "end",
                text=f"  👤  {full_name}",
                values=(reg,),
                tags=(f"emp_{emp.id}", "emp_node")
            )

        # 3. Recursively insert child departments (sorted alphabetically)
        if dept.children:
            sorted_children = sorted(dept.children, key=lambda d: str(d.name or "").lower())
            for child in sorted_children:
                self._insert_dept_node(iid, child)

    def _on_dept_select(self, event):
        # This is the legacy handler, now unified into _on_dept_select_hub
        # But we keep it for reference or internal state updates if needed
        selection = self.dept_tree.selection()
        if not selection: return
        item = selection[0]
        tags = self.dept_tree.item(item, "tags")
        if tags and "emp_node" in tags:
            self.selected_employee_id = int(tags[0].replace("emp_", ""))
            self._current_selected_emp_id = self.selected_employee_id
            self._update_photo_preview(self.selected_employee_id)
        else:
            self.selected_employee_id = None
            self._current_selected_emp_id = None
            self._update_photo_preview(None)
            if tags:
                self._selected_dept_id = int(tags[0])

    def _add_department(self):
        name = Querybox.get_string(prompt=tr("enter_department_name"), title=tr("add_department"), parent=self)
        if not name or not name.strip(): return
        try:
            self.service.create_department(name.strip(), self._selected_dept_id)
            self._refresh_dept_tree()
            Messagebox.show_info(tr("department_added", name=name), tr("success"), parent=self)
        except Exception as e: Messagebox.show_error(str(e), tr("error"), parent=self)

    def _rename_department(self):
        if not self._selected_dept_id:
            Messagebox.show_warning(tr("select_department_first"), tr("no_selection"), parent=self)
            return
        
        selection = self.dept_tree.selection()[0]
        curr_name = self.dept_tree.item(selection, "text").strip().replace("📁", "").strip()
        
        new_name = Querybox.get_string(prompt=tr("enter_new_name"), title=tr("rename_department"), initialvalue=curr_name, parent=self)
        if not new_name or not new_name.strip(): return
        
        try:
            from contragest.core.database import Department
            dept = self.session.query(Department).get(self._selected_dept_id)
            if dept:
                self.service.update_department(self._selected_dept_id, new_name.strip(), dept.parent_id)
                self._refresh_dept_tree()
                Messagebox.show_info(tr("department_renamed", old=curr_name, new=new_name.strip()), tr("success"), parent=self)
        except Exception as e: Messagebox.show_error(str(e), tr("error"), parent=self)

    def _delete_department(self):
        if not self._selected_dept_id:
            Messagebox.show_warning(tr("select_department_first"), tr("no_selection"), parent=self)
            return
        
        selection = self.dept_tree.selection()[0]
        dept_name = self.dept_tree.item(selection, "text").replace("  📁  ", "").strip()
        
        # 1. Security Verification (Dynamic Daily Password)
        correct_pwd = calculate_daily_password()
        pwd = Querybox.get_string(
            prompt=tr("enter_password"),
            title=tr("password_title"),
            parent=self
        )
        
        if pwd != correct_pwd:
            if pwd is not None:
                Messagebox.show_error(tr("incorrect_password"), tr("error"), parent=self)
            return
        
        # 2. Final Confirmation
        if Messagebox.okcancel(tr("confirm_delete_department", name=dept_name), tr("confirmation"), parent=self):
            try:
                self.service.delete_department(self._selected_dept_id)
                self._refresh_dept_tree()
                self._selected_dept_id = None
                Messagebox.show_info(tr("department_deleted", name=dept_name), tr("success"), parent=self)
            except Exception as e: 
                Messagebox.show_error(f"Deletion failed: {e}", tr("error"), parent=self)

    def new_contract_for_employee(self):
        """Open the Contract Form pre-linked to the selected employee."""
        if not self.selected_employee_id:
            Messagebox.show_warning(
                "Please select an employee from the table first.",
                tr("no_selection")
            )
            return
        from contragest.features.contracts.forms import ContractForm
        ContractForm(self, pre_employee_id=self.selected_employee_id)

    def search_employees(self):
        query = self.search_var.get().lower().strip()
        dept_val = self.dept_filter_var.get().lower().strip() if hasattr(self, "dept_filter_var") else ""
        role_val = self.role_filter_var.get().lower().strip() if hasattr(self, "role_filter_var") else ""
        
        if not query and not dept_val and not role_val:
            self.refresh_table_data()
            return
            
        filtered = []
        for row in self.all_employees_data:
            match_query = True
            if query and query != "search employees...":
                match_query = any(query in str(cell).lower() for cell in row.values())
                
            match_dept = True
            if dept_val:
                r_dept = row.get("department_name", "")
                match_dept = dept_val in str(r_dept).lower()
                
            match_role = True
            if role_val:
                r_role = row.get("role_title", "")
                match_role = role_val in str(r_role).lower()
                
            if match_query and match_dept and match_role:
                filtered.append(row)
                
        self.populate_table(filtered)

    def find_next(self):
        pass

    def import_data(self):
        """Open the Bulk Import window."""
        from contragest.features.employee_manager.bulk_import import BulkImportWindow
        BulkImportWindow(self, on_import_callback=self.load_all_employees)

    def update_table_columns(self):
        # Identify which columns are active
        active_config = [
            (label, attr) for label, attr, _ in self.field_config
            if self.column_vars.get(attr) and self.column_vars[attr].get()
        ]
        
        self.active_cols = [{"text": label, "stretch": True} for label, attr in active_config]
        self.active_attrs = [attr for label, attr in active_config]
        
        # Rebuild Tableview
        from ttkbootstrap.tableview import Tableview
        
        # Destroy previous table if it exists
        for widget in self.table_container.winfo_children():
            widget.destroy()
            
        self.emp_table = Tableview(
            master=self.table_container,
            coldata=self.active_cols,
            rowdata=[],
            paginated=True,
            pagesize=int(self.page_size_var.get()),
            searchable=False,
            bootstyle=PRIMARY,
            autofit=True
        )
        # Enable multi-selection (Ctrl/Shift+Click)
        self.emp_table.view.configure(selectmode="extended")
        self.emp_table.pack(fill=BOTH, expand=YES)
        self.emp_table.view.bind("<<TreeviewSelect>>", self.on_table_select)
        self.emp_table.view.bind("<Double-1>", self._on_double_click)
        self.emp_table.view.bind("<Button-3>", self._on_grid_right_click)
        
        # Reload data with current column selection
        self.refresh_table_data()

    def refresh_table_data(self):
        self.populate_table(self.all_employees_data)

    def populate_table(self, data_list):
        if not hasattr(self, "emp_table") or not self.emp_table:
            return
            
        # 1. High-Performance Persistent Sorting (Alphabetical & Numerical Alignment)
        # We sort by: Last Name (ASC) -> First Name (ASC) -> Reg Number (Numeric ASC)
        sorted_data_list = sorted(data_list, key=lambda x: (
            str(x.get("last_name", "")).strip().lower(),
            str(x.get("first_name", "")).strip().lower(),
            str(x.get("reg_number", "0")).zfill(12) # Robust numeric sort for strings
        ))
            
        row_data = []
        for emp_dict in sorted_data_list:
            row = []
            for attr in self.active_attrs:
                val = emp_dict.get(attr, "-")
                row.append(str(val) if val is not None else "-")
            row_data.append(row)
            
        try:
            # We use build_table_data to ensure columns and rows are perfectly aligned
            self.emp_table.build_table_data(self.active_cols, row_data)
            
            # 2. Visual Sort indicator: Default to Last Name or Reg Number if possible
            # Note: Tableview might reset internal sort indicators, so we enforce it here
            pass
        except Exception as e:
            logger.error(f"Table population error: {e}")
        
        if self.lbl_total_rows:
            self.lbl_total_rows.config(text=f"Total: {len(sorted_data_list)}")
        if self.lbl_selected_rows:
            self.lbl_selected_rows.config(text="Selected: 0")

    def load_all_employees(self):
        # Fetch ACTIVE employees (excluding archived ones)
        import time
        start_time = time.time()
        logger.info("Starting load_all_employees...")
        self.all_employees_data = []
        try:
            employees = self.service.get_all_active_employees()
            fetch_time = time.time() - start_time
            logger.info(f"Fetched {len(employees)} employees in {fetch_time:.2f}s")
            
            for e in employees:
                try:
                    e_dict = {
                        "selection_icon": "☐",
                        "id": e.id,
                        "registration_number": e.registration_number or "-",
                        "civility": e.civility or "-",
                        "last_name": e.last_name,
                        "first_name": e.first_name,
                        "department_name": e.dept_obj.name if e.dept_obj else (e.department or "Unassigned"),
                        "role_title": e.role_title or "-",
                        "matrimonial_status": e.matrimonial_status or "-",
                        "children_count": e.children_count or 0,
                        "cnss": e.cnss or "-",
                        "id_card_number": e.id_card_number or "-",
                        "passport": e.passport or "-",
                        "dob": e.dob.strftime("%Y-%m-%d") if e.dob else "-",
                        "nationality": e.nationality or "-",
                        "address": e.address or "-",
                        "mobile_phone": e.mobile_phone or "-",
                        "office_phone": e.office_phone or "-",
                        "privilege": e.privilege or "-",
                        "hire_date": e.hire_date.strftime("%Y-%m-%d") if e.hire_date else "-",
                        "exit_date": e.exit_date.strftime("%Y-%m-%d") if e.exit_date else "-",
                        "gross_salary": e.gross_salary or "-",
                        "net_salary": e.net_salary or "-",
                        "weekly_day_off": e.weekly_day_off or "NONE",
                        "is_auto_punch": "YES" if e.is_auto_punch else "NO",
                        "photo_icon": "🖼️" if e.photo_path and os.path.exists(e.photo_path) else "-"
                    }
                    self.all_employees_data.append(e_dict)
                except Exception as e_err:
                    logger.warning(f"Skipping employee due to data error: {e_err}")
            
            logger.info("Populating filter dropdowns...")
            self._populate_filter_dropdowns()
            
            logger.info("Refreshing table data...")
            self.refresh_table_data()
            
            total_time = time.time() - start_time
            logger.info(f"load_all_employees completed in {total_time:.2f}s")
            
        except Exception as e:
            logger.error(f"Error loading employees: {e}")
            import traceback
            logger.error(traceback.format_exc())
            Messagebox.show_error(f"Error loading employees: {e}", "Error")

    def _find_emp_id_from_row(self, item_values):
        """Helper to find Employee ID from table row values when ID column is hidden."""
        if not item_values or not self.active_attrs:
            return None
        # Values from tkinter Treeview are returned as a list, convert all to string for safe comparison
        str_item_values = [str(v) for v in item_values]
        
        for e in self.all_employees_data:
            match = True
            for i, attr in enumerate(self.active_attrs):
                # Skip the selection icon from matching logic as it's dynamic
                if attr == "selection_icon":
                    continue
                if i >= len(str_item_values):
                    match = False
                    break
                if str(e.get(attr, "-")) != str_item_values[i]:
                    match = False
                    break
            if match:
                return e.get("id")
        return None

    def on_table_select(self, event, manual_emp_id=None):
        if manual_emp_id:
            self._current_selected_emp_id = manual_emp_id
            self._update_photo_preview(manual_emp_id)
            return

        selection = self.emp_table.view.selection()
        
        # Update checkbox icons
        all_iids = self.emp_table.view.get_children()
        for iid in all_iids:
            icon = "☑" if iid in selection else "☐"
            self.emp_table.view.set(iid, 0, icon)

        if not selection: 
            self.selected_employee_ids = []
            self._current_selected_emp_id = None
            self._update_photo_preview(None)
            if self.lbl_selected_rows:
                self.lbl_selected_rows.config(text="Selected: 0")
            return
            
        self.selected_employee_ids = []
        for iid in selection:
            item = self.emp_table.view.item(iid)
            emp_id = self._find_emp_id_from_row(item.get('values', []))
            if emp_id:
                self.selected_employee_ids.append(emp_id)
        
        if self.lbl_selected_rows:
            self.lbl_selected_rows.config(text=f"Selected: {len(self.selected_employee_ids)}")
            
        if selection:
            item = self.emp_table.view.item(selection[0])
            first_emp_id = self._find_emp_id_from_row(item.get('values', []))
            self._current_selected_emp_id = first_emp_id
            self._update_photo_preview(first_emp_id)

        # Backward compatibility for single-selection logic
        self.selected_employee_id = self.selected_employee_ids[0] if self.selected_employee_ids else None
        self._update_photo_preview(self.selected_employee_id)

    def add_emp(self):
        """Open the Data Entry Form in add mode."""
        from contragest.features.employee_manager.data_entry_form import DataEntryForm
        DataEntryForm(self, mode="add", on_save_callback=self.load_all_employees)

    def save_employee(self):
        """Open the Data Entry Form in edit mode for the selected employee."""
        if not self.selected_employee_id:
            Messagebox.show_warning("Please select an employee from the table.", "No Selection")
            return
        from contragest.features.employee_manager.data_entry_form import DataEntryForm
        DataEntryForm(
            self, mode="edit",
            employee_id=self.selected_employee_id,
            on_save_callback=self.load_all_employees
        )

    def _on_double_click(self, event):
        """Handle double-click on a table row to open edit form."""
        if self.selected_employee_id:
            self.save_employee()

    def select_all_employees(self):
        """Select all employees in the current table view and update checkboxes."""
        self.emp_table.view.selection_set(self.emp_table.view.get_children())
        # The selection_set triggers the <<TreeviewSelect>> event which calls on_table_select
        # but just in case or for immediate feedback:
        self.on_table_select(None)

    def deselect_all_employees(self):
        """Clear selection and update checkboxes."""
        self.emp_table.view.selection_remove(self.emp_table.view.get_children())
        self.on_table_select(None)

    def archive_selected_employee(self):
        """Soft-delete selected employees (single or bulk) with reason and password."""
        if not hasattr(self, 'selected_employee_ids') or not self.selected_employee_ids:
            if not self.selected_employee_id:
                Messagebox.show_warning("Please select at least one employee.", "No Selection")
                return
            self.selected_employee_ids = [self.selected_employee_id]

        count = len(self.selected_employee_ids)
        
        # 1. Selection confirmation
        if count == 1:
            emp = self.service.get_employee(self.selected_employee_ids[0])
            prompt_name = f"'{emp.last_name} {emp.first_name}'" if emp else "this employee"
        else:
            prompt_name = f"{count} selected employees"

        # 2. Get Archive Reason (Once for the entire batch)
        reason = Querybox.get_string(
            f"Please provide a reason for archiving {prompt_name}:",
            "Archive Reason", parent=self
        )
        if not reason: return

        # 3. Password Check (Once for the entire batch)
        correct_pwd = calculate_daily_password()

        pwd_input = Querybox.get_string(
            prompt=tr("enter_password"),
            title="Authentication Required",
            parent=self,
            initialvalue=""
        )
        if pwd_input != correct_pwd:
            if pwd_input is not None:
                Messagebox.show_error(tr("incorrect_password"), tr("error"), parent=self)
            return

        # 4. Perform Archive
        success = self.service.bulk_archive_employees(self.selected_employee_ids, reason)
        if success:
            Messagebox.show_info(f"Successfully archived {count} employees.", "Success")
            self.load_all_employees()
            self.selected_employee_ids = []
            self.selected_employee_id = None
        else:
            Messagebox.show_error("Could not archive employees. Some records might be locked.", "Error")

    def open_archive_panel(self):
        """Prompt for password and open the Archive Management window."""
        correct_pwd = calculate_daily_password()

        pwd_input = Querybox.get_string(
            prompt=tr("enter_password"),
            title="Archive Access",
            parent=self,
            initialvalue=""
        )
        if pwd_input != correct_pwd:
            if pwd_input is not None:
                Messagebox.show_error(tr("incorrect_password"), tr("error"), parent=self)
            return

        from contragest.features.employee_manager.archive_panel import EmployeeArchivePanel
        EmployeeArchivePanel(self, on_reinstate_callback=self.load_all_employees)


    def delete_employee(self):
        if not self.selected_employee_id: 
            Messagebox.show_warning("Please select an employee.", "No Selection")
            return
            
        # 1. Security Verification (Dynamic Daily Password)
        correct_pwd = calculate_daily_password()
        pwd = Querybox.get_string(
            prompt=tr("enter_password"),
            title=tr("password_title"),
            parent=self
        )
        
        if pwd != correct_pwd:
            if pwd is not None:
                Messagebox.show_error(tr("incorrect_password"), tr("error"), parent=self)
            return

        # 2. Final Confirmation
        if Messagebox.okcancel("Confirm employee deletion? This action is IRREVERSIBLE.", "Delete"):
            self.service.delete_employee(self.selected_employee_id)
            self.load_all_employees()
            self.selected_employee_id = None

    def _get_export_data(self):
        """Helper to collect headers and rows for export based on active selection/filters."""
        # 1. Get active columns (excluding the selection checkbox)
        export_headers = [label for label, attr in zip(self.active_cols, self.active_attrs) if attr != "selection_icon"]
        header_texts = [h["text"] for h in export_headers]
        
        # 2. Get data rows
        # If rows are selected in the table, export only selected.
        # Otherwise, if table has data, export all visible rows.
        selection = self.emp_table.view.selection()
        
        row_data = []
        if selection:
            # Export selected rows
            for iid in selection:
                # Table data includes the checkbox column at index 0, we skip it
                values = self.emp_table.view.item(iid)["values"]
                row_data.append(values[1:])
        else:
            # Export all visible rows in the table
            for iid in self.emp_table.view.get_children():
                values = self.emp_table.view.item(iid)["values"]
                row_data.append(values[1:])
        
        return header_texts, row_data

    def export_selected_excel(self):
        headers, rows = self._get_export_data()
        if not rows:
            Messagebox.show_warning("No data selected or visible to export.", "Export Warning")
            return
            
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Export Employees to Excel",
            initialfile=f"Employees_Export_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )
        
        if not filepath:
            return
            
        self.status_bar.set_status("Exporting to Excel...", "loading")
        
        def run_export():
            try:
                export_utils.export_to_excel(rows, headers, filepath)
                self.after(0, lambda: Messagebox.show_info(f"Successfully exported data to:\n{filepath}", "Export Success"))
            except Exception as e:
                self.after(0, lambda e=e: Messagebox.show_error(f"Failed to export Excel:\n{str(e)}", "Export Error"))
            finally:
                self.after(0, lambda: self.status_bar.set_status("Ready", "idle"))
                
        threading.Thread(target=run_export, daemon=True).start()

    def export_selected_pdf(self):
        headers, rows = self._get_export_data()
        if not rows:
            Messagebox.show_warning("No data selected or visible to export.", "Export Warning")
            return
            
        filepath = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            title="Export Employees to PDF",
            initialfile=f"Employees_Export_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        )
        
        if not filepath:
            return
            
        self.status_bar.set_status("Exporting to PDF...", "loading")
        
        def run_export():
            try:
                export_utils.export_to_pdf(rows, headers, filepath)
                self.after(0, lambda: Messagebox.show_info(f"Successfully exported data to:\n{filepath}", "Export Success"))
            except Exception as e:
                self.after(0, lambda e=e: Messagebox.show_error(f"Failed to export PDF:\n{str(e)}", "Export Error"))
            finally:
                self.after(0, lambda: self.status_bar.set_status("Ready", "idle"))
                
        threading.Thread(target=run_export, daemon=True).start()

    def show_personal_sync_tab(self):
        """Switch to the Machine Sync/Personal tab."""
        if hasattr(self, 'personal_sync_container') and self.personal_sync_container:
            self.right_notebook.select(self.personal_sync_container)

    def _build_personal_sync_tab(self):
        parent = self.personal_sync_container
        self._personnel_machines = {} 

        # Machine selector
        sel_frame = ttk.Frame(parent)
        sel_frame.pack(fill=X, pady=3, padx=10)
        ttk.Label(sel_frame, text=f"🔌 {tr('machine_name')}:", font=("Inter", 9, "bold")).pack(side=LEFT, padx=3)

        self._personnel_machine_var = ttk.StringVar()
        self._personnel_combo = ttk.Combobox(sel_frame, textvariable=self._personnel_machine_var, width=40, state="readonly")
        self._personnel_combo.pack(side=LEFT, padx=3)

        ttk.Button(sel_frame, text="📡 " + tr("sync_employees").upper(), bootstyle=INFO, command=self._sync_personnel, width=20).pack(side=LEFT, padx=9)
        ttk.Button(sel_frame, text="🧬 DOWNLOAD BIOMETRICS", bootstyle="info-outline", command=self._sync_biometrics, width=24).pack(side=LEFT, padx=9)
        ttk.Button(sel_frame, text="🚀 UPLOAD BIOMETRICS", bootstyle="warning-outline", command=self._upload_biometrics, width=24).pack(side=LEFT, padx=9)

        # Action buttons
        btn_frame = ttk.Frame(parent)
        btn_frame.pack(fill=X, pady=3, padx=10)
        ttk.Button(btn_frame, text=f"📤 {tr('push_to_machine')}", bootstyle=SUCCESS, command=self._push_selected, padding=(7, 3)).pack(side=LEFT, padx=4)
        ttk.Button(btn_frame, text=f"🗑️ {tr('pull_from_machine')}", bootstyle=DANGER, command=self._pull_selected, padding=(7, 3)).pack(side=LEFT, padx=4)
        ttk.Button(btn_frame, text=f"🖼️ UPLOAD PHOTO", bootstyle=INFO, command=self._upload_photo_for_selected, padding=(7, 3)).pack(side=LEFT, padx=4)
        ttk.Button(btn_frame, text=f"📸 CAPTURE PHOTO", bootstyle="info-outline", command=self._capture_photo_for_selected, padding=(7, 3)).pack(side=LEFT, padx=4)
        ttk.Button(btn_frame, text=f"🗂️ BULK IMPORT", bootstyle="info-outline", command=self._bulk_import_photos_personal, padding=(7, 3)).pack(side=LEFT, padx=4)
        ttk.Button(btn_frame, text=f"📂 IMPORT BITMASKS", bootstyle="secondary-outline", command=self._import_bitmasks_from_file, padding=(7, 3)).pack(side=LEFT, padx=4)

        sep_btn = ttk.Separator(btn_frame, orient=VERTICAL)
        sep_btn.pack(side=LEFT, fill=Y, padx=6, pady=2)
        ttk.Button(btn_frame, text="📥 EXCEL", bootstyle="success-outline", command=self._export_selected_excel, padding=(7, 3)).pack(side=LEFT, padx=4)
        ttk.Button(btn_frame, text="📄 PDF", bootstyle="danger-outline", command=self._export_selected_pdf, padding=(7, 3)).pack(side=LEFT, padx=4)

        # Selection toolbar (lives outside the table frame so it survives syncs)
        sel_tool = ttk.Frame(parent)
        sel_tool.pack(fill=X, pady=(3, 0), padx=10)
        ttk.Button(sel_tool, text="☑️ SELECT ALL", bootstyle=INFO, command=self._select_all_personnel, padding=(6, 2)).pack(side=LEFT, padx=2)
        ttk.Button(sel_tool, text="⬜ DESELECT ALL", bootstyle=SECONDARY, command=self._deselect_all_personnel, padding=(6, 2)).pack(side=LEFT, padx=2)
        ttk.Button(sel_tool, text="🔄 INVERT", bootstyle="secondary-outline", command=self._invert_personnel_selection, padding=(6, 2)).pack(side=LEFT, padx=2)
        self._selection_count_lbl = ttk.Label(sel_tool, text="0 / 0 selected", font=("Inter", 8), foreground=TEXT_MUTED)
        self._selection_count_lbl.pack(side=LEFT, padx=8)

        # Container for split view
        self._personnel_paned = ttk.Panedwindow(parent, orient=HORIZONTAL)
        self._personnel_paned.pack(fill=BOTH, expand=YES, pady=3, padx=10)

        # Left: Table side
        self._personnel_table_frame = ttk.Frame(self._personnel_paned)
        self._personnel_paned.add(self._personnel_table_frame, weight=3)
        
        # Right: Tabbed Profile Card
        self._personnel_preview_frame = ttk.LabelFrame(self._personnel_paned, text="👤 PROFILE CARD")
        self._personnel_paned.add(self._personnel_preview_frame, weight=1)
        
        # Tabbed interface inside the card
        self._preview_notebook = ttk.Notebook(self._personnel_preview_frame, bootstyle="dark")
        self._preview_notebook.pack(fill=BOTH, expand=YES, padx=4, pady=4)
        
        # -- Tab 1: Photo
        self._tab_photo = ttk.Frame(self._preview_notebook)
        self._preview_notebook.add(self._tab_photo, text=" 🖼️  PHOTO ")
        
        self._photo_label = ttk.Label(self._tab_photo, text="SELECT AN EMPLOYEE\nTO VIEW DETAILS", anchor=CENTER, font=("Inter", 9), foreground=TEXT_MUTED)
        self._photo_label.pack(fill=BOTH, expand=YES, pady=20)
        self._detail_label = ttk.Label(self._tab_photo, text="", anchor=CENTER, justify=CENTER)
        self._detail_label.pack(fill=X, pady=6)

        # -- Tab 2: Biometrics
        self._tab_bio = ttk.Frame(self._preview_notebook)
        self._preview_notebook.add(self._tab_bio, text=" 🧬  BIOMETRICS ")
        
        self._bio_header = ttk.Label(self._tab_bio, text="No employee selected.", anchor=CENTER, font=("Inter", 8), foreground=TEXT_MUTED)
        self._bio_header.pack(fill=X, pady=6, padx=6)
        
        self._bio_canvas_frame = ttk.Frame(self._tab_bio)
        self._bio_canvas_frame.pack(fill=BOTH, expand=YES, padx=4)

        self._personnel_table = None

        # Reuse machine list
        self._refresh_personnel_combo()

    def _refresh_personnel_combo(self):
        machines = self.pointage_service.get_all_machines()
        self._personnel_machines = {f"{m.name} ({m.ip_address})": m.id for m in machines}
        self._personnel_combo["values"] = list(self._personnel_machines.keys())
        if self._personnel_machines:
            self._personnel_combo.current(0)

    def _get_selected_personnel_machine_id(self):
        key = self._personnel_machine_var.get()
        return self._personnel_machines.get(key)

    def _sync_personnel(self):
        mid = self._get_selected_personnel_machine_id()
        if not mid:
            Messagebox.show_info(tr("no_machine_selected"), tr("information"))
            return

        employees = self.pointage_service.get_employees_sync_status(mid)
        cols = ["ID", tr("reg_number") if tr("reg_number") != "reg_number" else "REG NUMBER",
                tr("employee"), tr("department"), tr("status")]
        rows = []
        for emp in employees:
            status = f"✅ {tr('on_machine')}" if emp["on_machine"] else f"❌ {tr('not_on_machine')}"
            rows.append((
                str(emp["id"]),
                str(emp.get("registration_number", "-")),
                str(emp["name"]),
                str(emp["department"]),
                status,
            ))

        for w in self._personnel_table_frame.winfo_children():
            w.destroy()

        from ttkbootstrap.tableview import Tableview
        self._personnel_table = Tableview(
            master=self._personnel_table_frame,
            coldata=cols,
            rowdata=rows,
            paginated=False,
            searchable=True,
            bootstyle="dark",
            autofit=True,
        )
        self._personnel_table.pack(fill=BOTH, expand=YES)
        
        # Bind selection
        self._personnel_table.view.bind("<<TreeviewSelect>>", self._on_personnel_select)
        self._personnel_table.view.bind("<Double-1>", self._on_personnel_double_click)
        self._on_personnel_select() # Refresh preview panel

    def _sync_biometrics(self):
        mid = self._get_selected_personnel_machine_id()
        if not mid:
            Messagebox.show_info(tr("no_machine_selected"), tr("information"))
            return
            
        self.status_bar.set_status("Scanning machine biometric registers...", "loading")
        
        def run_sync():
            try:
                machine = self.pointage_service.get_machine(mid)
                if not machine: return
                
                stats = self.pointage_service.connector.get_device_stats(
                    machine.ip_address, machine.port, machine.password or ""
                )
                face_count = stats.get("face_count", 0)
                face_cap   = stats.get("face_cap", 0)
                user_count = stats.get("user_count", 0)
                
                new_count = 0
                finger_count = stats.get('fingerprint_count', 0)
                if finger_count > 0:
                    new_count, total = self.pointage_service.sync_biometrics(mid)
                
                report = (
                    f"Biometric Scan Complete!\n\n"
                    f"Users on device:     {user_count}\n"
                    f"Face identities:     {face_count} / {face_cap}\n"
                    f"Finger templates:    {finger_count}\n"
                    f"Archived bitmasks:   {new_count} (new)\n\n"
                )
                if face_count > 0 and new_count == 0:
                    report += (
                        f"NOTE: Face bitmasks are stored in a secure\n"
                        f"firmware partition inaccessible via the\n"
                        f"standard network SDK (Port {machine.port}).\n"
                        f"Use 'UPLOAD PHOTO' to attach profile images."
                    )
                    
                self.after(0, lambda: Messagebox.show_info(report, "Biometric Report"))
            except Exception as e:
                self.after(0, lambda e=e: Messagebox.show_error(f"Scan failed: {e}", "Error"))
            finally:
                self.after(0, lambda: self.status_bar.set_status("Ready", "idle"))
                
        threading.Thread(target=run_sync, daemon=True).start()

    def _upload_biometrics(self):
        mid = self._get_selected_personnel_machine_id()
        if not mid:
            Messagebox.show_info(tr("no_machine_selected"), tr("information"))
            return
            
        confirm = Messagebox.show_question(
            "This will flash the archived Biometric Bitmasks (Faces/Fingers) from the database back into the physical machine memory.\n\nProceed with accelerated biometric upload?",
            "Confirm Upload"
        )
        if confirm != "Yes": return

        self.status_bar.set_status("Accelerated Bitmask Upload in progress...", "loading")
        
        def run_upload():
            try:
                success, failed = self.pointage_service.push_biometrics_to_machine(mid)
                self.after(0, lambda: Messagebox.show_info(
                    f"Upload Complete!\n\nSuccess: {success} bitmasks pushed.\nFailed: {failed} orphaned or invalid records.", 
                    "Success"
                ))
            except Exception as e:
                self.after(0, lambda e=e: self.status_bar.set_status(f"Error: {e}", "error"))
                self.after(0, lambda e=e: Messagebox.show_error(f"Upload failed: {e}", "Error"))
            finally:
                self.after(0, lambda: self.status_bar.set_status("Ready", "idle"))
                
        threading.Thread(target=run_upload, daemon=True).start()

    def _push_selected(self):
        if not self._personnel_table:
            return
        mid = self._get_selected_personnel_machine_id()
        if not mid:
            Messagebox.show_info(tr("no_machine_selected"), tr("information"))
            return
        sel = self._personnel_table.view.selection()
        if not sel:
            Messagebox.show_info(tr("no_employee_selected"), tr("information"))
            return
            
        success_count, failure_count = 0, 0
        
        for item_id in sel:
            values = self._personnel_table.view.item(item_id, "values")
            id_val = str(values[0])
            if id_val.startswith("ZK-"):
                failure_count += 1
                continue
                
            emp_id = int(id_val)
            if self.pointage_service.push_employee_to_machine(mid, emp_id):
                success_count += 1
            else:
                failure_count += 1
                
        self._sync_personnel()
        
        if failure_count == 0:
            Messagebox.show_info(f"Successfully pushed {success_count} user(s) to the machine.", tr("success"))
        elif success_count == 0:
            Messagebox.show_error(f"Failed to push {failure_count} user(s) to the machine.", tr("error"))
        else:
            Messagebox.show_warning(f"Partial success: {success_count} pushed, {failure_count} failed.", tr("information"))

    def _pull_selected(self):
        if not self._personnel_table:
            return
        mid = self._get_selected_personnel_machine_id()
        if not mid:
            Messagebox.show_info(tr("no_machine_selected"), tr("information"))
            return
        sel = self._personnel_table.view.selection()
        if not sel:
            Messagebox.show_info(tr("no_employee_selected"), tr("information"))
            return
            
        if Messagebox.okcancel(f"Are you sure you want to completely remove {len(sel)} user(s) from the physical time clock memory?", "Confirm Deletion") != "OK":
            return
            
        success_count, failure_count = 0, 0
        
        for item_id in sel:
            values = self._personnel_table.view.item(item_id, "values")
            id_val = str(values[0])
            status_val = str(values[4])
            
            is_not_on_machine = "not_on_machine" in status_val or "❌" in status_val
            
            if id_val.startswith("ZK-"):
                zk_uid = id_val.replace("ZK-", "")
                if is_not_on_machine:
                    result = True
                else:
                    result = self.pointage_service.remove_orphan_from_machine(mid, zk_uid)
            else:
                emp_id = int(id_val)
                if is_not_on_machine:
                    result = True
                else:
                    result = self.pointage_service.remove_employee_from_machine(mid, emp_id)
                
                if result:
                    from contragest.core.database import Employee
                    emp = self.session.query(Employee).get(emp_id)
                    if emp:
                        ans = Messagebox.show_question(
                            f"Employee '{emp.first_name} {emp.last_name}' (REG: {emp.registration_number}) is not on the machine.\n\n"
                            f"Would you also like to permanently delete this employee from the database?",
                            "Delete from Database",
                            buttons=["No:secondary", "Yes:danger"]
                        )
                        if ans == "Yes":
                            self.service.delete_employee(emp_id)
                
            if result:
                success_count += 1
            else:
                failure_count += 1
                
        self._sync_personnel()
        self.load_all_employees()
        
        if failure_count == 0:
            Messagebox.show_info(f"Successfully removed {success_count} user(s) from the machine.", tr("success"))
        elif success_count == 0:
            Messagebox.show_error(f"Failed to remove {failure_count} user(s) from the machine. Check connection.", tr("error"))
        else:
            Messagebox.show_warning(f"Partial success: {success_count} removed, {failure_count} failed.", tr("information"))

    def _upload_photo_for_selected(self):
        # Determine target employee
        emp_id = self.selected_employee_id
        emp_name = "Employee"
        
        if not emp_id and hasattr(self, "_personnel_table") and self._personnel_table:
            sel = self._personnel_table.view.selection()
            if sel:
                vals = self._personnel_table.view.item(sel[0], "values")
                if not str(vals[0]).startswith("ZK-"):
                    emp_id = int(vals[0])
                    emp_name = str(vals[2])

        if not emp_id:
            Messagebox.show_info("Please select an employee first.", "No Selection")
            return

        emp = self.service.get_employee_full(emp_id)
        if emp: emp_name = f"{emp.first_name} {emp.last_name}"
        
        path = filedialog.askopenfilename(
            title=f"Select Photo for {emp_name}",
            filetypes=[("Images", "*.png *.jpg *.jpeg *.bmp *.gif *.webp")]
        )
        if not path: return
            
        import shutil
        import os
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
        photos_dir = os.path.join(base_dir, "assets", "photos")
        os.makedirs(photos_dir, exist_ok=True)
        ext = os.path.splitext(path)[1]
        filename = f"emp_id{emp_id}{ext}"
        dest = os.path.join(photos_dir, filename)
        
        try:
            shutil.copy2(path, dest)
            self.service.update_employee(emp_id, photo_path=dest)
            self.status_bar.set_status(f"Photo uploaded for {emp_name}")
            self._update_photo_preview(emp_id)
            self.load_all_employees() # Refresh grid
            if hasattr(self, "_on_personnel_select"): self._on_personnel_select()
        except Exception as e:
            Messagebox.show_error(f"Failed to copy/save photo: {e}", tr("error"))

    def _capture_photo_for_selected(self):
        emp_id = self.selected_employee_id
        emp_name = "Employee"
        
        if not emp_id and hasattr(self, "_personnel_table") and self._personnel_table:
            sel = self._personnel_table.view.selection()
            if sel:
                vals = self._personnel_table.view.item(sel[0], "values")
                if not str(vals[0]).startswith("ZK-"):
                    emp_id = int(vals[0])
                    emp_name = str(vals[2])

        if not emp_id:
            Messagebox.show_info("Please select an employee first.", "No Selection")
            return

        emp = self.service.get_employee_full(emp_id)
        if emp: emp_name = f"{emp.first_name} {emp.last_name}"
        
        def on_capture(frame):
            import cv2
            import os
            base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
            photos_dir = os.path.join(base_dir, "assets", "photos")
            os.makedirs(photos_dir, exist_ok=True)
            filename = f"emp_id{emp_id}.jpg"
            dest = os.path.join(photos_dir, filename)
            
            try:
                cv2.imwrite(dest, frame, [int(cv2.IMWRITE_JPEG_QUALITY), 95])
                self.service.update_employee(emp_id, photo_path=dest)
                self.after(0, lambda: self.status_bar.set_status(f"Photo captured for {emp_name}"))
                self.after(0, lambda: self._update_photo_preview(emp_id))
                self.after(0, self.load_all_employees)
                if hasattr(self, "_on_personnel_select"): self.after(0, self._on_personnel_select)
            except Exception as e:
                self.after(0, lambda e=e: Messagebox.show_error(f"Capture Error: {e}", "Error"))

        WebcamCaptureDialog(self, emp_name, on_capture)

    def _bulk_import_photos_personal(self):
        import shutil
        from contragest.core.database import Employee

        dir_path = filedialog.askdirectory(title="Select Folder with Employee Photos", parent=self)
        if not dir_path:
            return

        valid_exts = {".jpg", ".jpeg", ".png", ".bmp", ".gif", ".webp"}
        
        files = []
        for root_dir, _, filenames in os.walk(dir_path):
            for f in filenames:
                if os.path.splitext(f)[1].lower() in valid_exts:
                    files.append(os.path.join(root_dir, f))
        
        if not files:
            Messagebox.show_warning("No valid images found in the selected folder or subfolders.", "Folder Empty", parent=self)
            return
            
        confirm = Messagebox.yesno(
            f"Found {len(files)} image(s) in the folder.\n\nThe system will match the filename (without extension) to the employee's REG NUMBER.\n\nProceed with Bulk Import?",
            "Confirm Bulk Import",
            parent=self
        )
        if confirm != "Yes":
            return
            
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
        photos_dir = os.path.join(base_dir, "assets", "photos")
        os.makedirs(photos_dir, exist_ok=True)
        
        success_count = 0
        ghost_count = 0
        error_count = 0

        prog = ttk.Toplevel(self)
        prog.title("Bulk Importing…")
        prog.geometry("380x120")
        prog.resizable(False, False)
        prog.transient(self)
        prog.grab_set()
        
        lbl = ttk.Label(prog, text="Mapping images to employees...", font=("Inter", 9, "bold"))
        lbl.pack(pady=(15, 5))
        pbar = ttk.Progressbar(prog, mode="determinate", bootstyle="info")
        pbar.pack(fill=X, padx=15, pady=5)
        self.update_idletasks()

        for i, src_path in enumerate(files):
            file_name = os.path.basename(src_path)
            
            pbar["value"] = int((i / len(files)) * 100)
            lbl.config(text=f"Processing {file_name}...")
            prog.update()
            
            basename, ext = os.path.splitext(file_name)
            raw_reg = str(basename).strip()
            
            clean_reg_num = raw_reg.lstrip('0') if raw_reg.isdigit() else raw_reg
            if not clean_reg_num and raw_reg.isdigit():
                clean_reg_num = "0"
            
            try:
                dest_filename = f"reg_{clean_reg_num}{ext}"
                dest_path = os.path.join(photos_dir, dest_filename)
                shutil.copy2(src_path, dest_path)

                from sqlalchemy import or_
                emp = self.session.query(Employee).filter(
                    or_(
                        Employee.registration_number == raw_reg,
                        Employee.registration_number == clean_reg_num
                    ),
                    Employee.is_archived == False
                ).first()
                
                if emp:
                    self.service.update_employee(emp.id, photo_path=dest_path)
                    success_count += 1
                else:
                    ghost_count += 1
                    
            except Exception as e:
                if error_count == 0:
                    self._first_error_msg_personal = str(e)
                error_count += 1

        prog.destroy()
        
        report = f"✅ DB Employees Updated: {success_count}\n"
        report += f"✅ Machine Only / Ghost Photos Saved: {ghost_count}\n"
        if error_count > 0:
            report += f"⚠️ Read/Write Errors: {error_count}\n"
            if hasattr(self, "_first_error_msg_personal"):
                report += f"\nFirst Error:\n{self._first_error_msg_personal}\n"
            
        Messagebox.show_info(report, "Bulk Import Complete", parent=self)
        
        if self._personnel_table:
            self._on_personnel_select()

    def _update_selection_count(self):
        """Refresh the 'N / M selected' label from the current table selection."""
        if not hasattr(self, "_selection_count_lbl"):
            return
        try:
            total = len(self._personnel_table.view.get_children()) if self._personnel_table else 0
            selected = len(self._personnel_table.view.selection()) if self._personnel_table else 0
            self._selection_count_lbl.configure(text=f"{selected} / {total} selected")
        except Exception:
            pass

    def _select_all_personnel(self):
        """Select every row in the personnel table."""
        if not self._personnel_table:
            return
        view = self._personnel_table.view
        items = view.get_children()
        if items:
            view.selection_set(*items)
        self._on_personnel_select()

    def _deselect_all_personnel(self):
        """Clear the personnel table selection."""
        if not self._personnel_table:
            return
        view = self._personnel_table.view
        if view.selection():
            view.selection_remove(*view.selection())
        self._on_personnel_select()

    def _invert_personnel_selection(self):
        """Invert the current personnel table selection."""
        if not self._personnel_table:
            return
        view = self._personnel_table.view
        current = set(view.selection())
        all_items = view.get_children()
        # The treeview's selection_* calls take items as positional args.
        if current:
            view.selection_remove(*current)
        new_items = [iid for iid in all_items if iid not in current]
        if new_items:
            view.selection_set(*new_items)
        self._on_personnel_select()

    def _on_personnel_select(self, event=None):
        if not self._personnel_table: return
        self._update_selection_count()
        sel = self._personnel_table.view.selection()
        if not sel: return
        
        values = self._personnel_table.view.item(sel[0], "values")
        id_val = str(values[0])
        
        if id_val.startswith("ZK-"):
            reg_num = str(values[1]).strip()
            photo_found = False
            base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
            for ext in ['.jpg', '.jpeg', '.png', '.bmp', '.gif', '.webp']:
                p = os.path.join(base_dir, "assets", "photos", f"reg_{reg_num}{ext}")
                if os.path.exists(p):
                    try:
                        img = Image.open(p)
                        img.thumbnail((200, 240))
                        self._current_prev_img = ImageTk.PhotoImage(img)
                        self._photo_label.configure(image=self._current_prev_img, text="")
                        photo_found = True
                        break
                    except: pass
            
            if not photo_found:
                self._photo_label.configure(image="", text="GHOST RECORD\n(Machine Only)")
                
            self._detail_label.configure(text=f"REG: {values[1]}\nNAME: {values[2]}\n\nThis user is not registered\nin the database.")
            self._bio_header.configure(text="No employee selected.", foreground=TEXT_MUTED)
            for w in self._bio_canvas_frame.winfo_children(): w.destroy()
            return

        emp_id = int(id_val)
        emp = self.service.get_employee_full(emp_id)
        
        if not emp:
            self._photo_label.configure(image="", text="ERROR\nLoading Profile")
            return

        info = f"NAME: {emp.first_name} {emp.last_name}\n"
        info += f"REG: {emp.registration_number}\n"
        info += f"DEPT: {emp.dept_obj.name if emp.dept_obj else (emp.department or '-')}\n"
        if emp.role_title: info += f"ROLE: {emp.role_title}"
        self._detail_label.configure(text=info, font=("Inter", 9, "bold"), foreground=DesignTokens.PRIMARY)

        if emp.photo_path and os.path.exists(emp.photo_path):
            try:
                img = Image.open(emp.photo_path)
                img.thumbnail((200, 240))
                self._current_prev_img = ImageTk.PhotoImage(img)
                self._photo_label.configure(image=self._current_prev_img, text="")
            except Exception:
                self._photo_label.configure(image="", text="INVALID\nIMAGE FILE")
        else:
            initials = f"{emp.first_name[0]}{emp.last_name[0]}".upper() if emp.first_name and emp.last_name else "?"
            self._photo_label.configure(
                image="",
                text=f"\n\n{initials}\n\nNO PHOTO\n\u2191 UPLOAD PHOTO",
                font=("Inter", 13, "bold"),
                foreground=DesignTokens.PRIMARY
            )

        if emp.registration_number:
            self._render_biometrics_panel(str(emp.registration_number))

    def _on_personnel_double_click(self, event):
        """Handle double-click to open employee edit form from the sync list."""
        if not self._personnel_table: return
        item = self._personnel_table.view.identify_row(event.y)
        if not item: return
        
        values = self._personnel_table.view.item(item, "values")
        id_val = str(values[0])
        
        if id_val.startswith("ZK-"):
            Messagebox.show_info(
                f"User {values[1]} is only on the machine.\n\nPlease 'PULL FROM MACHINE' first to create a database profile before editing.",
                "Ghost Record"
            )
            return

        self.selected_employee_id = int(id_val)
        self.save_employee()

    def _render_biometrics_panel(self, reg_num: str):
        for widget in self._bio_canvas_frame.winfo_children():
            widget.destroy()

        from contragest.core.database import BiometricTemplate
        templates = (
            self.session.query(BiometricTemplate)
            .filter_by(registration_number=reg_num)
            .order_by(BiometricTemplate.type, BiometricTemplate.template_index)
            .all()
        )

        if not templates:
            ttk.Label(
                self._bio_canvas_frame,
                text="No bitmask data archived\nfor this employee.\n\nUse 📂 IMPORT BITMASKS\nto load biometric data.",
                anchor=CENTER, justify=CENTER,
                font=("Inter", 8), foreground=TEXT_MUTED
            ).pack(expand=YES, pady=30)
            self._bio_header.configure(text=f"Bitmasks: 0  |  REG: {reg_num}")
            return

        self._bio_header.configure(
            text=f"Bitmasks: {len(templates)}  |  REG: {reg_num}",
            foreground=DesignTokens.PRIMARY
        )

        import ast
        TYPE_ICONS = {"face": "👁️", "finger": "🖐️"}
        TYPE_COLORS = {"face": DesignTokens.PRIMARY, "finger": DesignTokens.SUCCESS}

        for t in templates:
            row = ttk.Frame(self._bio_canvas_frame)
            row.pack(fill=X, pady=2, padx=2)

            icon = TYPE_ICONS.get(t.type, "⬛")
            color = TYPE_COLORS.get(t.type, TEXT_MUTED)

            info_frame = ttk.Frame(row)
            info_frame.pack(side=LEFT, padx=4, pady=4)
            ttk.Label(info_frame, text=icon, font=("Segoe UI Emoji", 14)).pack()
            ttk.Label(info_frame, text=f"{t.type.upper()}\nSlot {t.template_index}", font=("Inter", 7), foreground=color, justify=CENTER).pack()

            canvas_frame = ttk.Frame(row)
            canvas_frame.pack(side=LEFT, fill=BOTH, expand=YES, padx=4, pady=4)

            try:
                raw = ast.literal_eval(t.template_data)
                byte_vals = list(raw[:64]) if isinstance(raw, (bytes, bytearray)) else []
            except Exception:
                try:
                    hex_str = t.template_data.strip()
                    byte_vals = [int(hex_str[i:i+2], 16) for i in range(0, min(128, len(hex_str)), 2)]
                except Exception:
                    byte_vals = []

            size_label = f"{len(byte_vals)} B" if byte_vals else "-"
            ttk.Label(canvas_frame, text=f"v{t.version} · {size_label}", font=("Inter", 7), foreground=TEXT_MUTED).pack(anchor=W)

            c = tk.Canvas(canvas_frame, height=14, bg="#0d0d0d", highlightthickness=0)
            c.pack(fill=X, pady=2)
            c.update_idletasks()
            w = max(c.winfo_width(), 160)
            bw = max(1, w // max(len(byte_vals), 1)) if byte_vals else 4
            for i, bval in enumerate(byte_vals[:64]):
                hue = int((bval / 255) * 200)
                r = min(255, hue * 2)
                g = min(255, (200 - hue))
                b = 180
                color_hex = f"#{r:02x}{g:02x}{b:02x}"
                c.create_rectangle(i * bw, 0, i * bw + bw, 14, fill=color_hex, outline="")

    def _import_bitmasks_from_file(self):
        path = filedialog.askopenfilename(
            title="Import Biometric Bitmasks",
            filetypes=[("CSV Files", "*.csv"), ("JSON Files", "*.json"), ("All Files", "*.*")]
        )
        if not path: return

        self.status_bar.set_status("Importing Biometric Bitmasks...", "loading")

        def run_import():
            import json, csv
            from contragest.core.database import BiometricTemplate

            records = []
            errors = []

            try:
                if path.lower().endswith(".json"):
                    with open(path, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    if isinstance(data, dict): data = [data]
                    records = data
                else:
                    with open(path, "r", encoding="utf-8") as f:
                        reader = csv.DictReader(f)
                        records = list(reader)
            except Exception as e:
                self.after(0, lambda e=e: Messagebox.show_error(f"Failed to read file: {e}", "Import Error"))
                self.after(0, lambda: self.status_bar.set_status("Ready", "idle"))
                return

            REQUIRED = {"registration_number", "type", "template_index", "template_data"}
            imported = 0
            skipped = 0

            for idx, rec in enumerate(records):
                missing = REQUIRED - set(rec.keys())
                if missing:
                    errors.append(f"Row {idx+1}: missing fields {missing}")
                    skipped += 1
                    continue

                btype = str(rec["type"]).lower().strip()
                if btype not in ("finger", "face"):
                    errors.append(f"Row {idx+1}: invalid type '{btype}' (must be 'finger' or 'face')")
                    skipped += 1
                    continue

                tdata = str(rec["template_data"]).strip()
                if not tdata:
                    errors.append(f"Row {idx+1}: empty template_data")
                    skipped += 1
                    continue

                try:
                    tidx = int(rec["template_index"])
                    version = int(rec.get("version", 10))
                except (ValueError, TypeError) as e:
                    errors.append(f"Row {idx+1}: {e}")
                    skipped += 1
                    continue

                reg = str(rec["registration_number"]).strip()
                if not reg:
                    errors.append(f"Row {idx+1}: empty registration_number")
                    skipped += 1
                    continue

                existing = self.session.query(BiometricTemplate).filter_by(
                    registration_number=reg, type=btype, template_index=tidx
                ).first()

                if existing:
                    existing.template_data = tdata
                    existing.version = version
                else:
                    self.session.add(BiometricTemplate(
                        registration_number=reg,
                        type=btype,
                        template_index=tidx,
                        template_data=tdata,
                        version=version
                    ))
                imported += 1

            try:
                self.session.commit()
            except Exception as e:
                self.session.rollback()
                self.after(0, lambda e=e: Messagebox.show_error(f"DB commit failed: {e}", "Import Error"))
                self.after(0, lambda: self.status_bar.set_status("Ready", "idle"))
                return

            report = f"Import Complete!\n\nImported: {imported}\nSkipped:  {skipped}\nTotal rows: {len(records)}"
            if errors[:5]:
                report += "\n\nFirst errors:\n" + "\n".join(errors[:5])

            self.after(0, lambda: Messagebox.show_info(report, "Bitmask Import"))
            self.after(0, lambda: self.status_bar.set_status("Ready", "idle"))
            self.after(0, self._on_personnel_select)

        threading.Thread(target=run_import, daemon=True).start()

    def _get_personnel_export_rows(self):
        """Return selected table rows, or all rows if none selected.

        Each row is a dict with keys: id, reg_number, employee, department,
        status.  Falls back gracefully when the table is empty.
        """
        if not self._personnel_table:
            return []
        view = self._personnel_table.view
        sel = view.selection()
        items = sel if sel else view.get_children()
        rows = []
        for iid in items:
            vals = view.item(iid, "values")
            if len(vals) < 5:
                continue
            rows.append({
                "id": str(vals[0]),
                "reg_number": str(vals[1]),
                "employee": str(vals[2]),
                "department": str(vals[3]),
                "status": str(vals[4]),
            })
        return rows

    def _export_selected_excel(self):
        """Export selected (or all) employees to an Excel workbook."""
        rows = self._get_personnel_export_rows()
        if not rows:
            Messagebox.show_info("No data to export.", "Export", parent=self)
            return

        default_name = f"employees_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("All files", "*.*")],
            initialfile=default_name,
            title="Export to Excel",
            parent=self,
        )
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Employees"

            headers = ["ID", "Reg Number", "Employee", "Department", "Status"]
            col_widths = [8, 14, 30, 22, 24]

            header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color=getattr(DesignTokens, "SECONDARY", "#1E293B").lstrip("#"),
                                     end_color=getattr(DesignTokens, "SECONDARY", "#1E293B").lstrip("#"),
                                     fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"),
            )

            for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
                ws.column_dimensions[get_column_letter(col_idx)].width = w

            data_font = Font(name="Calibri", size=10)
            for r_idx, row in enumerate(rows, 2):
                for c_idx, key in enumerate(["id", "reg_number", "employee", "department", "status"], 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=row[key])
                    cell.font = data_font
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")

            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
            ws.freeze_panes = "A2"

            wb.save(path)
            Messagebox.show_info(f"Exported {len(rows)} employees to Excel.\n{path}", "Export Success", parent=self)
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            Messagebox.show_error(f"Export failed: {e}", "Export Error", parent=self)

    def _export_selected_pdf(self):
        """Export selected (or all) employees to a PDF report."""
        rows = self._get_personnel_export_rows()
        if not rows:
            Messagebox.show_info("No data to export.", "Export", parent=self)
            return

        default_name = f"employees_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf"), ("All files", "*.*")],
            initialfile=default_name,
            title="Export to PDF",
            parent=self,
        )
        if not path:
            return

        try:
            class EmployeePDF(FPDF):
                def header(self):
                    self.set_font("Helvetica", "B", 14)
                    self.cell(0, 10, "Employee Report", align="C", new_x="LMARGIN", new_y="NEXT")
                    self.set_font("Helvetica", "", 9)
                    self.cell(0, 6, datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                              align="C", new_x="LMARGIN", new_y="NEXT")
                    self.ln(4)

                def footer(self):
                    self.set_y(-15)
                    self.set_font("Helvetica", "I", 8)
                    self.cell(0, 10, f"Page {self.page_no()}/{{nb}}", align="C")

            pdf = EmployeePDF(orientation="L", unit="mm", format="A4")
            pdf.alias_nb_pages()
            pdf.add_page()

            headers = ["ID", "Reg Number", "Employee", "Department", "Status"]
            col_widths = [18, 30, 80, 55, 60]

            pdf.set_font("Helvetica", "B", 10)
            pdf.set_fill_color(30, 41, 59)
            pdf.set_text_color(255, 255, 255)
            for h, w in zip(headers, col_widths):
                pdf.cell(w, 8, h, border=1, align="C", fill=True)
            pdf.ln()

            pdf.set_font("Helvetica", "", 9)
            pdf.set_text_color(0, 0, 0)
            fill = False
            for row in rows:
                if fill:
                    pdf.set_fill_color(241, 245, 249)
                else:
                    pdf.set_fill_color(255, 255, 255)
                for key, w in zip(["id", "reg_number", "employee", "department", "status"], col_widths):
                    pdf.cell(w, 7, str(row[key]), border=1, fill=True)
                pdf.ln()
                fill = not fill

            pdf.output(path)
            Messagebox.show_info(f"Exported {len(rows)} employees to PDF.\n{path}", "Export Success", parent=self)
        except Exception as e:
            logger.error(f"PDF export failed: {e}")
            Messagebox.show_error(f"Export failed: {e}", "Export Error", parent=self)

    def destroy(self):
        try:
            if hasattr(self, 'session') and self.session:
                self.session.close()
        except Exception:
            pass
        super().destroy()

    def __del__(self):
        try:
            if hasattr(self, 'session') and self.session:
                self.session.close()
        except: pass
