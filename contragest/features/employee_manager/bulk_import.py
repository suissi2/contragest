"""
Bulk Import Window - CSV/Excel import with preview, column mapping, and validation.

Workflow:
  1. User selects a CSV or XLSX file.
  2. Preview grid displays parsed rows.
  3. Column mapping auto-detects standard names; user can adjust.
  4. Validation summary highlights errors per row.
  5. Import commits valid rows and reports skipped ones.
"""

import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from ttkbootstrap.dialogs import Messagebox
from ttkbootstrap.tableview import Tableview
from contragest.core.database import SessionLocal
from contragest.features.employee_manager.service import EmployeeService
from contragest.core.i18n import tr
from tkinter import filedialog
import csv
import os


# Standard column name mappings (file header → DB field)
COLUMN_ALIASES = {
    "first_name": ["first_name", "firstname", "first name", "prénom", "prenom"],
    "last_name": ["last_name", "lastname", "last name", "nom", "nom de famille"],
    "email": ["email", "e-mail", "mail", "courriel"],
    "department": ["department", "dept", "département", "departement", "service"],
    "function": ["function", "job", "job_title", "fonction", "poste"],
    "role_title": ["role", "role_title", "titre"],
    "civility": ["civility", "civilité", "civilite", "title", "mr/ms"],
    "registration_number": ["registration_number", "reg_num", "matricule", "reg"],
    "mobile_phone": ["mobile", "mobile_phone", "tel_mobile", "téléphone mobile"],
    "office_phone": ["office_phone", "phone", "tel", "téléphone"],
    "nationality": ["nationality", "nationalité", "nationalite"],
    "address": ["address", "adresse"],
    "dob": ["dob", "date_of_birth", "birth_date", "date de naissance"],
    "hire_date": ["hire_date", "date_hired", "date d'embauche", "date embauche"],
    "id_card_number": ["id_card", "id_card_number", "cin", "carte identité"],
    "passport": ["passport", "passeport"],
    "cnss": ["cnss", "social_security"],
    "gross_salary": ["gross_salary", "salaire_brut", "salaire brut"],
    "net_salary": ["net_salary", "salaire_net", "salaire net"],
    "matrimonial_status": ["matrimonial_status", "marital_status", "situation familiale"],
    "children_count": ["children_count", "children", "enfants", "nombre enfants"],
    "privilege": ["privilege", "privilège"],
}

# Required fields for import
REQUIRED_FIELDS = {"first_name", "last_name"}


class BulkImportWindow(ttk.Toplevel):
    """Window for importing employee data from CSV/Excel files."""

    def __init__(self, parent, on_import_callback=None):
        super().__init__(parent)
        self.title(tr("bulk_import_title"))
        self.geometry("1000x700")
        self.resizable(True, True)
        self.grab_set()

        self.on_import_callback = on_import_callback
        self.session = SessionLocal()
        self.service = EmployeeService(self.session)

        self.raw_headers = []
        self.parsed_rows = []
        self.column_mapping = {}  # file_header → db_field

        self._build_ui()
        self.center_window()

    # ------------------------------------------------------------------ #
    #  UI
    # ------------------------------------------------------------------ #

    def _build_ui(self):
        # Header
        header = ttk.Frame(self, bootstyle="warning")
        header.pack(fill=X)
        ttk.Label(
            header, text=f"  📥  {tr('bulk_import_title')}",
            font=("Helvetica", 13, "bold"),
            bootstyle="inverse-warning", padding=(15, 8)
        ).pack(side=LEFT)

        # File selection row
        file_frame = ttk.Frame(self, padding=10)
        file_frame.pack(fill=X)

        ttk.Label(file_frame, text=tr("select_file") + ":").pack(side=LEFT)
        self.file_var = ttk.StringVar()
        ttk.Entry(file_frame, textvariable=self.file_var, state="readonly", width=60).pack(side=LEFT, padx=10, fill=X, expand=YES)

        ttk.Button(
            file_frame, text=f"📂 {tr('browse')}", bootstyle="info",
            command=self._pick_file
        ).pack(side=LEFT, padx=5)

        # Column mapping panel (collapsible)
        mapping_frame = ttk.LabelFrame(self, text=tr("column_mapping"))
        mapping_frame.pack(fill=X, padx=10, pady=5)

        self.mapping_container = ttk.Frame(mapping_frame)
        self.mapping_container.pack(fill=X)

        # Preview table
        preview_label = ttk.LabelFrame(self, text=tr("preview"))
        preview_label.pack(fill=BOTH, expand=YES, padx=10, pady=5)

        self.table_container = ttk.Frame(preview_label)
        self.table_container.pack(fill=BOTH, expand=YES)

        # Status / validation summary
        self.status_var = ttk.StringVar(value=tr("no_file_selected"))
        ttk.Label(self, textvariable=self.status_var, font=("Helvetica", 10)).pack(fill=X, padx=10, pady=5)

        # Action bar
        action_bar = ttk.Frame(self, padding=10)
        action_bar.pack(fill=X, side=BOTTOM)

        ttk.Button(
            action_bar, text=f"❌ {tr('cancel')}", bootstyle="secondary-outline",
            command=self.destroy, padding=(20, 8)
        ).pack(side=RIGHT, padx=5)

        self.import_btn = ttk.Button(
            action_bar, text=f"📥 {tr('import_now')}", bootstyle="success",
            command=self._do_import, padding=(20, 8), state=DISABLED
        )
        self.import_btn.pack(side=RIGHT, padx=5)

    # ------------------------------------------------------------------ #
    #  File parsing
    # ------------------------------------------------------------------ #

    def _pick_file(self):
        """Open file dialog and parse selected file."""
        path = filedialog.askopenfilename(
            title=tr("select_file"),
            filetypes=[
                ("CSV files", "*.csv"),
                ("Excel files", "*.xlsx"),
                ("All files", "*.*")
            ]
        )
        if not path:
            return

        self.file_var.set(path)
        ext = os.path.splitext(path)[1].lower()

        try:
            if ext == ".csv":
                self._parse_csv(path)
            elif ext == ".xlsx":
                self._parse_xlsx(path)
            else:
                Messagebox.show_warning(tr("unsupported_format"), tr("error"))
                return
        except Exception as e:
            Messagebox.show_error(f"{tr('parse_error')}: {str(e)}", tr("error"))
            return

        self._auto_map_columns()
        self._build_mapping_ui()
        self._show_preview()
        self.import_btn.configure(state=NORMAL)

    def _parse_csv(self, path):
        """Parse a CSV file into headers + row dicts."""
        with open(path, 'r', encoding='utf-8-sig') as f:
            # Detect delimiter
            sample = f.read(4096)
            f.seek(0)
            sniffer = csv.Sniffer()
            try:
                dialect = sniffer.sniff(sample, delimiters=',;\t|')
            except csv.Error:
                dialect = csv.excel

            reader = csv.DictReader(f, dialect=dialect)
            self.raw_headers = reader.fieldnames or []
            self.parsed_rows = [row for row in reader]

        self.status_var.set(tr("rows_loaded", count=len(self.parsed_rows)))

    def _parse_xlsx(self, path):
        """Parse an Excel file (requires openpyxl)."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            Messagebox.show_error(
                tr("openpyxl_missing"),
                tr("error")
            )
            return

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            self.raw_headers = []
            self.parsed_rows = []
            return

        self.raw_headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
        self.parsed_rows = []
        for row in rows[1:]:
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(self.raw_headers):
                    row_dict[self.raw_headers[i]] = str(val).strip() if val is not None else ""
            self.parsed_rows.append(row_dict)
        wb.close()

        self.status_var.set(tr("rows_loaded", count=len(self.parsed_rows)))

    # ------------------------------------------------------------------ #
    #  Column mapping
    # ------------------------------------------------------------------ #

    def _auto_map_columns(self):
        """Automatically map file headers to DB fields using alias matching."""
        self.column_mapping = {}
        for header in self.raw_headers:
            h_lower = header.lower().strip()
            mapped = None
            for db_field, aliases in COLUMN_ALIASES.items():
                if h_lower in aliases:
                    mapped = db_field
                    break
            self.column_mapping[header] = mapped  # None = unmapped

    def _build_mapping_ui(self):
        """Build combobox mapping UI for each file column."""
        for widget in self.mapping_container.winfo_children():
            widget.destroy()

        db_field_options = ["-- skip --"] + list(COLUMN_ALIASES.keys())
        self.mapping_vars = {}

        # Use grid layout for compact display
        cols = 3  # Number of mapping pairs per row
        for idx, header in enumerate(self.raw_headers):
            row = idx // cols
            col = (idx % cols) * 2

            ttk.Label(
                self.mapping_container, text=header,
                font=("Helvetica", 9), foreground="gray"
            ).grid(row=row, column=col, padx=5, pady=2, sticky=W)

            var = ttk.StringVar(value=self.column_mapping.get(header) or "-- skip --")
            self.mapping_vars[header] = var
            combo = ttk.Combobox(
                self.mapping_container, textvariable=var,
                values=db_field_options, width=18, state="readonly"
            )
            combo.grid(row=row, column=col + 1, padx=5, pady=2)

    # ------------------------------------------------------------------ #
    #  Preview
    # ------------------------------------------------------------------ #

    def _show_preview(self):
        """Show the first 50 rows in a preview table."""
        for widget in self.table_container.winfo_children():
            widget.destroy()

        if not self.parsed_rows:
            return

        col_data = [{"text": h, "stretch": True} for h in self.raw_headers]
        row_data = []
        for row in self.parsed_rows[:50]:
            row_data.append([row.get(h, "") for h in self.raw_headers])

        table = Tableview(
            master=self.table_container,
            coldata=col_data,
            rowdata=row_data,
            paginated=True,
            pagesize=15,
            searchable=False,
            bootstyle=INFO,
            autofit=True
        )
        table.pack(fill=BOTH, expand=YES)

    # ------------------------------------------------------------------ #
    #  Import
    # ------------------------------------------------------------------ #

    def _get_final_mapping(self):
        """Read the mapping comboboxes and return {file_header: db_field} (excluding skips)."""
        mapping = {}
        for header, var in self.mapping_vars.items():
            val = var.get()
            if val and val != "-- skip --":
                mapping[header] = val
        return mapping

    def _do_import(self):
        """Validate & import all rows."""
        mapping = self._get_final_mapping()

        # Validate at least required fields are mapped
        mapped_fields = set(mapping.values())
        missing_required = REQUIRED_FIELDS - mapped_fields
        if missing_required:
            Messagebox.show_error(
                tr("missing_required_mapping", fields=", ".join(missing_required)),
                tr("error")
            )
            return

        # Convert rows using mapping
        records = []
        for row in self.parsed_rows:
            record = {}
            for file_header, db_field in mapping.items():
                val = row.get(file_header, "").strip()
                if val:
                    record[db_field] = val
            records.append(record)

        # Bulk import via service
        try:
            success_count, errors = self.service.bulk_import_employees(records)
            summary = tr("import_summary",
                         success=success_count,
                         errors=len(errors),
                         total=len(records))
            if errors:
                summary += "\n\n" + tr("import_errors_detail") + ":\n"
                summary += "\n".join(errors[:20])  # Show max 20 errors
                if len(errors) > 20:
                    summary += f"\n... +{len(errors) - 20} more"

            Messagebox.show_info(summary, tr("import_complete"))

            if self.on_import_callback and success_count > 0:
                self.on_import_callback()
            self.destroy()

        except Exception as e:
            Messagebox.show_error(str(e), tr("error"))

    # ------------------------------------------------------------------ #
    #  Utilities
    # ------------------------------------------------------------------ #

    def center_window(self):
        self.update_idletasks()
        w = self.winfo_width()
        h = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (w // 2)
        y = (self.winfo_screenheight() // 2) - (h // 2)
        self.geometry(f"{w}x{h}+{x}+{y}")

    def destroy(self):
        try:
            if hasattr(self, 'session'):
                self.session.close()
        except Exception:
            pass
        super().destroy()
