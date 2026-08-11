import os
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from fpdf import FPDF
from contragest.core.database import SessionLocal, AppConfig
from contragest.core.i18n import tr
from contragest.core.logging import setup_logger
from contragest.core.pdf_utils import safe_pdf_str
from contragest.core.excel_utils import safe_excel_sheet_title, safe_excel_cell_str

logger = setup_logger("export_reports")

def _safe_str(val, default=""):
    """Heavily defensive string conversion for Excel/PDF exports."""
    if pd.isna(val) or val is None:
        return default
    return safe_excel_cell_str(str(val).strip())

def _safe_time_slice(val, length=5):
    """Safely slices a time string (e.g. HH:MM:SS -> HH:MM)."""
    s = _safe_str(val)
    if not s or s == "-":
        return "-"
    return s[:length]

def _get_company_logo_path():
    session = SessionLocal()
    try:
        config = session.query(AppConfig).first()
        if config and config.company_logo_path and os.path.exists(config.company_logo_path):
            return config.company_logo_path
    except Exception:
        pass
    finally:
        session.close()
    return None

def time_to_timedelta(t_str):
    if not t_str or t_str == "-":
        return pd.Timedelta(seconds=0)
    try:
        parts = t_str.split(':')
        if len(parts) == 2:
            return pd.Timedelta(hours=int(parts[0]), minutes=int(parts[1]))
        elif len(parts) == 3:
            return pd.Timedelta(hours=int(parts[0]), minutes=int(parts[1]), seconds=int(parts[2]))
    except:
        pass
    return pd.Timedelta(seconds=0)

def format_timedelta(td):
    total_seconds = int(td.total_seconds())
    hours, remainder = divmod(abs(total_seconds), 3600)
    minutes, _ = divmod(remainder, 60)
    sign = "-" if total_seconds < 0 else ""
    return f"{sign}{hours:02d}:{minutes:02d}"

def format_date_locale(date_str, short=False):
    if not date_str:
        return ""
    try:
        dt = datetime.strptime(date_str, '%Y-%m-%d')
        from contragest.core.i18n import get_lang_manager
        lang = get_lang_manager().current_lang
        if lang == 'fr':
            months_fr = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin', 'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
            days_fr = ['Lun.', 'Mar.', 'Mer.', 'Jeu.', 'Ven.', 'Sam.', 'Dim.']
            day_name = days_fr[dt.weekday()]
            month_name = months_fr[dt.month - 1]
            if short:
                return f"{day_name} {dt.day:02d}-{dt.month:02d}-{dt.year}"
            return f"{day_name} {dt.day} {month_name} {dt.year}"
        else:
            if short:
                return dt.strftime('%a. %d-%m-%Y')
            return dt.strftime('%a. %d %B %Y')

    except:
        return date_str

def generate_daily_attendance_excel(data, filepath, from_date=None, to_date=None):
    if not data:
        raise ValueError("No data available to export.")
        
    wb = Workbook()
    ws = wb.active
    
    # Title
    date_range_str = ""
    f_date = format_date_locale(from_date) if from_date else ""
    t_date = format_date_locale(to_date) if to_date else ""
    
    if from_date and to_date and from_date != to_date:
        date_range_str = f"{f_date} To {t_date}"
    elif from_date:
        date_range_str = f"{f_date}"
    elif to_date:
        date_range_str = f"{t_date}"
        
    ws.title = safe_excel_sheet_title(tr("Daily Attendance Report"))

    
    # Add Logo
    logo_path = _get_company_logo_path()
    if logo_path:
        try:
            img = Image(logo_path)
            aspect_ratio = img.width / img.height
            img.height = 60
            img.width = 60 * aspect_ratio
            ws.add_image(img, 'A1')
            ws.row_dimensions[1].height = 60
        except Exception:
            pass

    title_cell = ws.cell(row=1, column=1)
    full_title = f"Daily Attendance Report {date_range_str}".strip()
    title_cell.value = full_title.upper()

    title_cell.font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    headers = [
        tr("DATE"), tr("DEPARTMENT"), tr("MAT."), tr("LAST & FIRST NAME"), 
        tr("STATUS"), tr("CHECK IN 1"), tr("CHECK OUT 1"), tr("CHECK IN 2"), 
        tr("CHECK OUT 2"), tr("ATTENDANCE TIME"), tr("WORK TIME"), tr("DIFFERENCE"), tr("NOTE")
    ]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    alignment_center = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin', color="BFBFBF"), 
                         right=Side(style='thin', color="BFBFBF"), 
                         top=Side(style='thin', color="BFBFBF"), 
                         bottom=Side(style='thin', color="BFBFBF"))
                         
    row_idx = 3
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = safe_excel_cell_str(header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        
    row_idx += 1
    
    df = pd.DataFrame(data)
    if df.empty:
        try:
            wb.save(filepath)
        except Exception as e:
            raise RuntimeError(f"Could not save Excel file. Is it already open? Error: {e}")
        return
        
    sort_key = 'raw_date' if 'raw_date' in df.columns else 'date'
    df.sort_values(by=[sort_key, 'department', 'employee'], inplace=True)
    if 'id' in df.columns:
        # Synthetic (Cartesian-padded) rows use id=-1 or None; only rows with
        # a positive integer id are real DB rows that need deduplication.
        mask_real = df['id'].notna() & (df['id'] != -1)
        df_real = df[mask_real].drop_duplicates(subset=['id'])
        df_synth = df[~mask_real]
        df = pd.concat([df_real, df_synth], ignore_index=True)
        df.sort_values(by=[sort_key, 'department', 'employee'], inplace=True)
    elif not df.empty:
        df.drop_duplicates(inplace=True)

    
    grouped_date = df.groupby(sort_key, sort=False)
    
    group_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    dept_fill = PatternFill(start_color="EAEFFF", end_color="EAEFFF", fill_type="solid")
    total_fill = PatternFill(start_color="ACB9CA", end_color="ACB9CA", fill_type="solid")
    bold_font = Font(bold=True)
    
    grand_total_jours = 0
    grand_total_travail = pd.Timedelta(seconds=0)
    grand_total_presence = pd.Timedelta(seconds=0)
    grand_status = {}

    for date_key, date_group in grouped_date:
        date_travail = pd.Timedelta(seconds=0)
        date_presence = pd.Timedelta(seconds=0)
        date_jours_presents = 0
        date_status = {}
        
        grouped_dept = date_group.groupby('department', sort=False)
        for dept_key, dept_group in grouped_dept:
            dept_travail = pd.Timedelta(seconds=0)
            dept_presence = pd.Timedelta(seconds=0)
            dept_jours_presents = 0
            dept_status = {}
            
            for _, row in dept_group.iterrows():
                try:
                    date_val = format_date_locale(_safe_str(row.get('date')), short=True)
                    st = _safe_str(row.get('status'))
                    if st:
                        dept_status[st] = dept_status.get(st, 0) + 1
                        date_status[st] = date_status.get(st, 0) + 1
                        grand_status[st] = grand_status.get(st, 0) + 1
                    
                    ci1 = _safe_str(row.get('check_in'), "-")
                    co1 = _safe_str(row.get('check_out'), "-")
                    ci2 = _safe_str(row.get('check_in_2'), "-")
                    co2 = _safe_str(row.get('check_out_2'), "-")
                    
                    at = _safe_str(row.get('attendance_time'), "-")
                    wt = _safe_str(row.get('work_time'), "-")
                    
                    ws.append([
                        date_val, _safe_str(row.get('department')), _safe_str(row.get('reg_number')), _safe_str(row.get('employee')),
                        st, ci1, co1, ci2, co2, at, wt, _safe_str(row.get('difference'), "-"), _safe_str(row.get('note'))
                    ])
                    
                    current_row = ws.max_row
                    for c in range(1, 14):
                        cell = ws.cell(row=current_row, column=c)
                        cell.alignment = alignment_center
                        cell.border = thin_border
                        
                    if at and at != "-":
                        dept_jours_presents += 1
                        dept_presence += time_to_timedelta(at)
                        date_presence += time_to_timedelta(at)
                        grand_total_presence += time_to_timedelta(at)
                    if wt and wt != "-":
                        dept_travail += time_to_timedelta(wt)
                        date_travail += time_to_timedelta(wt)
                        grand_total_travail += time_to_timedelta(wt)
                except Exception as e:
                    logger.error(f"Error processing row for Excel export: {e}")
                    # Continue to next row instead of crashing the whole export
                    continue
                    
            # SUB-TOTAL for Department
            date_jours_presents += dept_jours_presents
            grand_total_jours += dept_jours_presents
            
            # Format status breakdown (excluding 'P' as it's represented by dept_jours_presents)
            status_parts = []
            for st, count in dept_status.items():
                if st != 'P': status_parts.append(f"{st}:{count}")
            st_summary = " | ".join(status_parts) if status_parts else ""

            ws.append(["", "", "", f"{tr('Total')} {dept_key}", dept_jours_presents, "", "", "", "", 
                       format_timedelta(dept_presence), format_timedelta(dept_travail), "", st_summary])
            
            last_row = ws.max_row
            for c in range(1, 14):
                cell = ws.cell(row=last_row, column=c)
                cell.font = bold_font
                cell.fill = dept_fill
                cell.border = thin_border
                cell.alignment = alignment_center
            
            for st, count in dept_status.items():
                if st == 'P' or count == 0: continue
                ws.append(["", "", "", f"{tr('Total')} {st}", count, "", "", "", "", "", "", "", ""])
                lr = ws.max_row
                for c in range(1, 14):
                    cell = ws.cell(row=lr, column=c)
                    cell.fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
                    cell.border = thin_border
                    cell.font = bold_font
                    if c == 4: cell.alignment = Alignment(horizontal='right')
                    else: cell.alignment = alignment_center

        # DATE LEVEL TOTAL
        group_date_str = format_date_locale(date_key)
        ws.append(["", "", "", f"{tr('Total pour')} {group_date_str}", date_jours_presents, "", "", "", "", 
                   format_timedelta(date_presence), format_timedelta(date_travail), "", ""])
        lr = ws.max_row
        for c in range(1, 14):
            cell = ws.cell(row=lr, column=c)
            cell.font = bold_font
            cell.fill = group_fill
            cell.border = thin_border
            cell.alignment = alignment_center
            
        # Only show date-level status breakdown if there's more than one day
        if len(grouped_date) > 1:
            for st, count in date_status.items():
                if st == 'P' or count == 0: continue
                ws.append(["", "", "", f"{tr('Total')} {st}", count, "", "", "", "", "", "", "", ""])
                lr = ws.max_row
                for c in range(1, 14):
                    cell = ws.cell(row=lr, column=c)
                    cell.fill = PatternFill(start_color="EAECEF", end_color="EAECEF", fill_type="solid")
                    cell.border = thin_border
                    cell.font = bold_font
                    if c == 4: cell.alignment = Alignment(horizontal='right')
                    else: cell.alignment = alignment_center
        ws.append([]) # Blank spacer


    # GRAND TOTAL
    ws.append([tr("Total Jours Présents"), "", "", "", grand_total_jours, "", "", "", "", 
               format_timedelta(grand_total_presence), format_timedelta(grand_total_travail), "", ""])
    lr = ws.max_row
    ws.merge_cells(start_row=lr, start_column=1, end_row=lr, end_column=4)
    for c in range(1, 14):
        cell = ws.cell(row=lr, column=c)
        cell.font = bold_font
        cell.fill = total_fill
        cell.border = thin_border
        cell.alignment = alignment_center
        
    for st, count in grand_status.items():
        if st == 'P' or count == 0: continue
        ws.append([f"{tr('Total')} {st}", "", "", "", count, "", "", "", "", "", "", "", ""])
        lr = ws.max_row
        ws.merge_cells(start_row=lr, start_column=1, end_row=lr, end_column=4)
        for c in range(1, 14):
            cell = ws.cell(row=lr, column=c)
            cell.font = bold_font
            cell.fill = total_fill
            cell.border = thin_border
            if c == 1: cell.alignment = Alignment(horizontal='right')
            else: cell.alignment = alignment_center

    column_widths = {'A': 18, 'B': 22, 'C': 12, 'D': 30, 'E': 12, 'F': 14, 'G': 14, 'H': 14, 'I': 14, 'J': 15, 'K': 15, 'L': 15, 'M': 25}
    for col, width in column_widths.items(): ws.column_dimensions[col].width = width
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0
    ws.page_margins.right = 0
    ws.page_margins.header = 0
    ws.page_margins.footer = 0
    ws.print_title_rows = '1:3'
    ws.oddFooter.center.text = "&P / &N"
    ws.evenFooter.center.text = "&P / &N"
    try:
        wb.save(filepath)
    except Exception as e:
        raise RuntimeError(f"Could not save Excel file. Is it already open? Error: {e}")

def generate_attendance_excel(data, filepath, from_date=None, to_date=None):
    if not data:
        raise ValueError("No data available to export.")
        
    wb = Workbook()
    ws = wb.active
    
    f_date = format_date_locale(from_date) if from_date else ""
    t_date = format_date_locale(to_date) if to_date else ""
    if from_date and to_date:
        if from_date == to_date:
            date_range_str = f"{f_date}"
        else:
            date_range_str = f"{f_date} To {t_date}"
    elif from_date:
        date_range_str = f"{f_date}"
    elif to_date:
        date_range_str = f"{t_date}"
    else:
        date_range_str = ""
        
    ws.title = safe_excel_sheet_title(tr("Attendance Report"))

    logo_path = _get_company_logo_path()
    if logo_path:
        try:
            img = Image(logo_path)
            img.height = 60
            img.width = 60 * (img.width/img.height)
            ws.add_image(img, 'A1')
            ws.row_dimensions[1].height = 60
        except: pass
            
    title_cell = ws.cell(row=1, column=1, value=f"Attendance Report {date_range_str}".strip().upper())

    title_cell.font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    headers = [tr("DATE"), tr("DEPARTMENT"), tr("MAT."), tr("LAST & FIRST NAME"), tr("STATUS"), tr("CHECK IN 1"), tr("CHECK OUT 1"), tr("CHECK IN 2"), tr("CHECK OUT 2"), tr("ATTENDANCE TIME"), tr("WORK TIME"), tr("DIFFERENCE"), tr("NOTE")]
    h_font, h_fill = Font(bold=True, color="FFFFFF"), PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(left=Side(style='thin', color="BFBFBF"), right=Side(style='thin', color="BFBFBF"), top=Side(style='thin', color="BFBFBF"), bottom=Side(style='thin', color="BFBFBF"))
    
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c_idx, value=h)
        cell.font, cell.fill, cell.border = h_font, h_fill, thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    df = pd.DataFrame(data)
    if df.empty:
        try:
            wb.save(filepath)
        except Exception as e:
            raise RuntimeError(f"Could not save Excel file. Is it already open? Error: {e}")
        return
    
    sort_date_key = 'raw_date' if 'raw_date' in df.columns else 'date'
    df.sort_values(by=['department', 'reg_number', sort_date_key], inplace=True)
    if 'id' in df.columns:
        # Synthetic (Cartesian-padded) rows use id=-1 or None; only rows with
        # a positive integer id are real DB rows that need deduplication.
        mask_real = df['id'].notna() & (df['id'] != -1)
        df_real = df[mask_real].drop_duplicates(subset=['id'])
        df_synth = df[~mask_real]
        df = pd.concat([df_real, df_synth], ignore_index=True)
        df.sort_values(by=['department', 'reg_number', sort_date_key], inplace=True)
    elif not df.empty:
        df.drop_duplicates(inplace=True)

    # Use a unique grouping key (Name + Reg) to avoid collisions between employees with same name
    df['group_key'] = df['employee'].astype(str) + " (" + df['reg_number'].astype(str) + ")"
    grouped = df.groupby('group_key', sort=False)

    
    group_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_fill = PatternFill(start_color="ACB9CA", end_color="ACB9CA", fill_type="solid")
    bold_font, alignment_center = Font(bold=True), Alignment(horizontal='center', vertical='center')
    
    # Rely on service-provided gap filling
    grand_total_jours, grand_total_travail, grand_total_presence = 0, pd.Timedelta(seconds=0), pd.Timedelta(seconds=0)
    grand_total_diff = pd.Timedelta(seconds=0)  # signed: positive=overtime, negative=deficit
    grand_status = {}


    for employee, group in grouped:
        total_travail, total_presence, total_diff, jours_presents = pd.Timedelta(seconds=0), pd.Timedelta(seconds=0), pd.Timedelta(seconds=0), 0
        
        # Sort each employee's records by date (using raw_date if available)
        d_sort_key = 'raw_date' if 'raw_date' in group.columns else 'date'
        group = group.sort_values(by=d_sort_key)
        
        for _, row in group.iterrows():
            try:
                date_val = _safe_str(row.get('date'))
                # Handle ISO date format (YYYY-MM-DD) vs locale format (Sun. DD-MM-YY)
                if len(date_val) == 10 and date_val.count('-') == 2 and not date_val.endswith(' '):
                    try:
                        date_val = format_date_locale(date_val, short=True)
                    except: pass
                    
                ci1 = _safe_str(row.get('check_in'), "-")
                co1 = _safe_str(row.get('check_out'), "-")
                ci2 = _safe_str(row.get('check_in_2'), "-")
                co2 = _safe_str(row.get('check_out_2'), "-")
                at, wt = _safe_str(row.get('attendance_time'), "-"), _safe_str(row.get('work_time'), "-")
                
                data_row = [date_val, _safe_str(row.get('department')), _safe_str(row.get('reg_number')), _safe_str(row.get('employee')), _safe_str(row.get('status'), "-"), ci1, co1, ci2, co2, at, wt, _safe_str(row.get('difference')), _safe_str(row.get('note'))]
                ws.append(data_row)
                
                # Style the row efficiently
                for cell in ws[ws.max_row]:
                    cell.alignment = alignment_center
                    cell.border = thin_border
                
                if at and at != "-":
                    jours_presents += 1
                    try: total_presence += time_to_timedelta(at)
                    except: pass
                    
                st = _safe_str(row.get('status'))
                if st and st != 'nan' and st != '-':
                    grand_status[st] = grand_status.get(st, 0) + 1

                if wt and wt != "-":
                    try: total_travail += time_to_timedelta(wt)
                    except: pass

                # Accumulate signed difference (e.g. "+01:00" or "-00:13")
                diff_raw = _safe_str(row.get('difference'), "")
                if diff_raw and diff_raw not in ("-", ""):
                    try:
                        sign = -1 if diff_raw.startswith("-") else 1
                        diff_clean = diff_raw.lstrip("+-")
                        total_diff += sign * time_to_timedelta(diff_clean)
                    except Exception:
                        pass
            except Exception as e:
                logger.error(f"Error processing row in period export: {e}")
                continue

        # Subtotal — includes signed DIFFERENCE sum in column 12
        diff_sign = "+" if total_diff.total_seconds() >= 0 else "-"
        diff_abs = abs(total_diff)
        diff_h, diff_rem = divmod(int(diff_abs.total_seconds()), 3600)
        diff_m = diff_rem // 60
        total_diff_str = f"{diff_sign}{diff_h:02d}:{diff_m:02d}" if total_diff.total_seconds() != 0 else "+00:00"
        ws.append(["", "", "", tr("Jours Présents"), jours_presents, "", "", "", "", format_timedelta(total_presence), format_timedelta(total_travail), total_diff_str, ""])
        lr = ws.max_row
        for c in range(1, 14):
            cell = ws.cell(row=lr, column=c)
            cell.border, cell.fill, cell.font, cell.alignment = thin_border, group_fill, bold_font, alignment_center
        
        # ADD EMPLOYEE LEVEL STATUS BREAKDOWN (like in daily report)
        emp_status = group['status'].value_counts()
        for st_name, count in emp_status.items():
            st_name = str(st_name).strip()
            if not st_name or st_name == 'P' or st_name == 'nan': continue
            ws.append(["", "", "", f"{tr('Total')} {st_name}", count, "", "", "", "", "", "", "", ""])
            lr = ws.max_row
            for c in range(1, 14):
                cell = ws.cell(row=lr, column=c)
                cell.fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
                cell.border = thin_border
                cell.font = bold_font
                if c == 4: cell.alignment = Alignment(horizontal='right')
                else: cell.alignment = alignment_center

        
        grand_total_jours += jours_presents
        grand_total_travail += total_travail
        grand_total_presence += total_presence
        grand_total_diff += total_diff
        
    # Grand Totals (with overall DIFFERENCE sum in column 12)
    g_diff_sign = "+" if grand_total_diff.total_seconds() >= 0 else "-"
    g_diff_abs = abs(grand_total_diff)
    g_diff_h, g_diff_rem = divmod(int(g_diff_abs.total_seconds()), 3600)
    g_diff_m = g_diff_rem // 60
    grand_diff_str = f"{g_diff_sign}{g_diff_h:02d}:{g_diff_m:02d}" if grand_total_diff.total_seconds() != 0 else "+00:00"
    ws.append([tr("Total Jours Présents"), "", "", "", grand_total_jours, "", "", "", "", format_timedelta(grand_total_presence), format_timedelta(grand_total_travail), grand_diff_str, ""])
    lr = ws.max_row
    ws.merge_cells(start_row=lr, start_column=1, end_row=lr, end_column=4)
    for c in range(1, 14):
        cell = ws.cell(row=lr, column=c)
        cell.font, cell.fill, cell.border, cell.alignment = bold_font, total_fill, thin_border, alignment_center
    
    # Add Status Breakdown to Period Report
    row_idx = lr + 1
    for st, count in grand_status.items():
        if st == 'P' or count == 0: continue
        ws.append([f"{tr('Total')} {st}", "", "", "", count, "", "", "", "", "", "", "", ""])
        lr = ws.max_row
        ws.merge_cells(start_row=lr, start_column=1, end_row=lr, end_column=4)
        for c in range(1, 14):
            cell = ws.cell(row=lr, column=c)
            cell.font, cell.fill, cell.border = bold_font, total_fill, thin_border
            if c == 1: cell.alignment = Alignment(horizontal='right')
            else: cell.alignment = alignment_center

    
    col_widths = {'A': 18, 'B': 22, 'C': 12, 'D': 30, 'E': 12, 'F': 14, 'G': 14, 'H': 14, 'I': 14, 'J': 15, 'K': 15, 'L': 15, 'M': 25}
    for col, width in col_widths.items(): ws.column_dimensions[col].width = width
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0
    ws.page_margins.right = 0
    ws.page_margins.header = 0
    ws.page_margins.footer = 0
    ws.print_title_rows = '1:3'
    ws.oddFooter.center.text = "&P / &N"
    ws.evenFooter.center.text = "&P / &N"
    
    # Second sheet: RECAP BY DEPARTMENT
    ws2 = wb.create_sheet(title=safe_excel_sheet_title(tr("RECAP BY DEPARTMENT")))
    recap_cols = [tr('Departement'), tr('Present'), 'P', 'AB', 'CA', 'JF', 'RH', 'RHB', 'CR', 'CM', 'MAP', 'PJF', 'MIS', 'JFP', 'CSS', 'DS']
    for s in df['status'].unique():
        sc = str(s).strip()
        if sc and sc not in recap_cols: recap_cols.append(sc)
    recap_cols.append(tr('Total'))
    
    recap_data = {}
    for _, row in df.iterrows():
        dept = row.get('department', 'Unknown') or 'Unknown'
        if dept not in recap_data: recap_data[dept] = {c: 0 for c in recap_cols if c != tr('Departement')}
        st = str(row.get('status', '')).strip()
        at = row.get('attendance_time', '-')
        if st and st in recap_data[dept]: recap_data[dept][st] += 1
        elif at and at != "-": recap_data[dept][tr('Present')] += 1
            
    for dept in recap_data:
        recap_data[dept][tr('Total')] = sum(recap_data[dept][c] for c in recap_cols if c not in (tr('Departement'), tr('Total')))
    
    b_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    h_fill_r, c_fill_r = PatternFill(start_color="E7E6E6", fill_type="solid"), PatternFill(start_color="D9D9D9", fill_type="solid")

    for col_idx, h in enumerate(recap_cols, 1):
        cell = ws2.cell(row=5, column=col_idx, value=h)
        cell.font, cell.fill, cell.border = bold_font, h_fill_r, b_border
    
    row_idx = 6
    g_totals = {c: 0 for c in recap_cols if c != tr('Departement')}
    for dept in sorted(recap_data.keys()):
        ws2.cell(row=row_idx, column=1, value=dept).border = b_border
        ws2.cell(row=row_idx, column=1).fill = c_fill_r
        for col_idx, col_name in enumerate(recap_cols[1:], 2):
            val = recap_data[dept][col_name]
            cell = ws2.cell(row=row_idx, column=col_idx, value=val if val > 0 else "")
            cell.border, cell.fill = b_border, c_fill_r
            g_totals[col_name] += val
        row_idx += 1
        
    cell_tot = ws2.cell(row=row_idx, column=1, value=tr("Total"))
    cell_tot.font, cell_tot.border, cell_tot.fill = bold_font, b_border, h_fill_r
    for col_idx, col_name in enumerate(recap_cols[1:], 2):
        val = g_totals[col_name]
        cell = ws2.cell(row=row_idx, column=col_idx, value=val if val > 0 else "")
        cell.font, cell.border, cell.fill = bold_font, b_border, h_fill_r
        
    ws2.column_dimensions['A'].width = 25
    try:
        wb.save(filepath)
    except Exception as e:
        raise RuntimeError(f"Could not save Excel file. Is it already open? Error: {e}")

def generate_attendance_pdf(data, filepath, from_date=None, to_date=None):
    if not data:
        raise ValueError("No data available to export.")
    df = pd.DataFrame(data)
    
    class PDF(FPDF):
        def footer(self):
            self.set_y(-15); self.set_font("helvetica", "I", 8)
            self.cell(0, 10, f"Page {self.page_no()}", align="C")

    pdf = PDF(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    
    logo_path = _get_company_logo_path()
    if logo_path:
        try: pdf.image(logo_path, x=10, y=8, h=36)
        except: pass
            
    title_x = 35 if logo_path else 10
    pdf.set_xy(title_x, 10); pdf.set_font("helvetica", "B", 16)
    f_date = format_date_locale(from_date); t_date = format_date_locale(to_date)
    date_range_str = f"{tr('from {from_date} to {to_date}', from_date=f_date, to_date=t_date)}" if from_date and to_date else ""
    pdf.cell(0, 8, safe_pdf_str(f"{tr('Attendance Summary Report')} {date_range_str}").strip(), align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("helvetica", "I", 10); pdf.cell(0, 8, safe_pdf_str(f"{tr('Generated:')} {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"), align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(15)

    pdf.set_font("helvetica", "B", 12); pdf.cell(0, 10, safe_pdf_str(tr("RÉSUMÉ PAR DÉPARTEMENT")), new_x="LMARGIN", new_y="NEXT"); pdf.ln(2)
    u_emps = df.drop_duplicates(subset=['reg_number'])
    dept_counts = u_emps['department'].value_counts().reset_index()
    dept_counts.columns = [tr('Departement'), tr('Total')]
    t_emps = dept_counts[tr('Total')].sum()
    start_y = pdf.get_y()
    
    pdf.set_font("helvetica", "B", 10); pdf.set_fill_color(31, 78, 121); pdf.set_text_color(255, 255, 255)
    pdf.cell(20, 8, safe_pdf_str(tr("Total")), border=1, align="C", fill=True); pdf.cell(70, 8, safe_pdf_str(tr("Departement")), border=1, align="C", fill=True, new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("helvetica", "", 9); pdf.set_text_color(0, 0, 0)
    
    fill = False
    for _, row in dept_counts.iterrows():
        pdf.set_fill_color(240, 240, 240) if fill else pdf.set_fill_color(255, 255, 255)
        pdf.cell(20, 7, safe_pdf_str(str(row[tr('Total')])), border=1, align="C", fill=True); pdf.cell(70, 7, safe_pdf_str(str(row[tr('Departement')])), border=1, align="L", fill=True, new_x="LMARGIN", new_y="NEXT")
        fill = not fill
        
    pdf.set_font("helvetica", "B", 9); pdf.set_fill_color(220, 230, 241)
    pdf.cell(20, 7, safe_pdf_str(str(t_emps)), border=1, align="C", fill=True); pdf.cell(70, 7, safe_pdf_str(tr("Total")), border=1, align="L", fill=True, new_x="LMARGIN", new_y="NEXT")
    
    chart_x, chart_y, chart_w, chart_h = 120, start_y, 160, 100
    pdf.set_xy(chart_x, chart_y); pdf.set_font("helvetica", "", 8); pdf.rect(chart_x, chart_y, chart_w, chart_h, style='D')
    pdf.set_xy(chart_x, chart_y + 2); pdf.cell(chart_w, 5, safe_pdf_str(tr("Répartition Par Département (Funnel)")), align="C", new_x="LMARGIN", new_y="NEXT")
    
    if not dept_counts.empty:
        max_v = dept_counts[tr('Total')].max()
        bar_x, bar_y, bar_w, bar_h = chart_x + 10, chart_y + 15, chart_w - 20, chart_h - 45
        for i in range(6): pdf.line(bar_x, bar_y + bar_h - (i * bar_h / 5), bar_x + bar_w, bar_y + bar_h - (i * bar_h / 5))
        b_w = min(10, (bar_w - 10) / len(dept_counts)) if len(dept_counts) > 0 else 10
        colors = [(31, 119, 180), (255, 127, 14), (44, 160, 44), (214, 39, 40), (148, 103, 189), (140, 86, 75), (227, 119, 194), (127, 127, 127), (188, 189, 34), (23, 190, 207)]
        leg_y, leg_x = bar_y + bar_h + 10, bar_x
        for idx, (_, row) in enumerate(dept_counts.iterrows()):
            h = (row[tr('Total')] / max_v) * bar_h if max_v > 0 else 0
            pdf.set_fill_color(*colors[idx % len(colors)])
            pdf.rect(bar_x + 5 + idx * (b_w + 2), bar_y + bar_h - h, b_w, h, style='F')
            if leg_x > chart_x + chart_w - 40: leg_x, leg_y = bar_x, leg_y + 6
            pdf.rect(leg_x, leg_y, 4, 4, style='F'); pdf.set_xy(leg_x + 5, leg_y); pdf.set_text_color(100, 100, 100); pdf.set_font("helvetica", "B", 6)
            lt = str(row[tr('Departement')]); pdf.cell(30, 4, safe_pdf_str(lt[:17] + "..." if len(lt) > 20 else lt))
            leg_x += 40
            
    pdf.add_page("P"); pdf.set_text_color(0, 0, 0); pdf.set_font("helvetica", "B", 12); pdf.cell(0, 10, safe_pdf_str(tr("Temps Moyen de Présence par Employé")), new_x="LMARGIN", new_y="NEXT"); pdf.ln(2)
    e_stats = []
    for emp, group in df.groupby('employee'):
        tot_p, days_p = pd.Timedelta(seconds=0), 0
        reg, dep = group.iloc[0].get('reg_number', ''), group.iloc[0].get('department', '')
        for _, row in group.iterrows():
            at = row.get('attendance_time', '-')
            if at and at != "-":
                tot_p += time_to_timedelta(at); days_p += 1
        if days_p > 0:
            avg_p = tot_p / days_p
            e_stats.append({tr('No ID'): reg, tr('Nom'): emp, tr('Departement'): dep, tr('Tot. T. De Prés.'): format_timedelta(tot_p), tr('Moy. T. De Prés.'): format_timedelta(avg_p), '_avg': avg_p.total_seconds()})
    
    e_stats.sort(key=lambda x: x['_avg'], reverse=True)
    pdf.set_font("helvetica", "B", 9); pdf.set_fill_color(31, 78, 121); pdf.set_text_color(255, 255, 255)
    cols = [(tr("No ID"), 20), (tr("Nom"), 60), (tr("Departement"), 50), (tr("Tot. T. De Prés."), 30), (tr("Moy. T. De Prés."), 30)]
    for n, w in cols: pdf.cell(w, 8, safe_pdf_str(n), border=1, align="C", fill=True)
    pdf.ln(); pdf.set_font("helvetica", "", 8); pdf.set_text_color(0, 0, 0)
    fill = False
    for r in e_stats:
        try:
            pdf.set_fill_color(240, 240, 240) if fill else pdf.set_fill_color(255, 255, 255)
            pdf.cell(cols[0][1], 7, safe_pdf_str(str(r[tr('No ID')])), border=1, align="C", fill=True)
            pdf.cell(cols[1][1], 7, safe_pdf_str(r[tr('Nom')][:27] + "..." if len(r[tr('Nom')]) > 30 else r[tr('Nom')]), border=1, align="L", fill=True)
            pdf.cell(cols[2][1], 7, safe_pdf_str(r[tr('Departement')][:22] + "..." if len(r[tr('Departement')]) > 25 else r[tr('Departement')]), border=1, align="L", fill=True)
            t, m = str(r[tr('Tot. T. De Prés.')]), str(r[tr('Moy. T. De Prés.')])
            pdf.cell(cols[3][1], 7, safe_pdf_str(t[:-3] if t.endswith(':00') else t), border=1, align="C", fill=True)
            pdf.cell(cols[4][1], 7, safe_pdf_str(m[:-3] if m.endswith(':00') else m), border=1, align="C", fill=True)
            pdf.ln()
            fill = not fill
        except Exception as e:
            logger.error(f"Error processing employee stat row for PDF: {e}")
            continue
    try:
        pdf.output(filepath)
    except Exception as e:
        raise RuntimeError(f"Could not save PDF file. Is it already open? Error: {e}")

def generate_detailed_attendance_pdf(data, filepath, from_date=None, to_date=None):
    if not data:
        raise ValueError("No data available to export.")
    
    df = pd.DataFrame(data)
    
    # Deduplication logic (same as Excel)
    sort_key = 'raw_date' if 'raw_date' in df.columns else 'date'
    if 'id' in df.columns:
        mask_real = df['id'].notna() & (df['id'] != -1)
        df_real = df[mask_real].drop_duplicates(subset=['id'])
        df_synth = df[~mask_real]
        df = pd.concat([df_real, df_synth], ignore_index=True)

    # Sort for effective grouping: Dept -> Employee Name -> Date
    df.sort_values(by=['department', 'employee', sort_key], inplace=True)

    class PDF(FPDF):
        def footer(self):
            self.set_y(-15)
            self.set_font("helvetica", "I", 8)
            self.cell(0, 10, f"Page {self.page_no()}", align="C")

    pdf = PDF(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    
    # Header logic
    logo_path = _get_company_logo_path()
    if logo_path:
        try: pdf.image(logo_path, x=10, y=8, h=20)
        except: pass
            
    pdf.set_xy(10, 10)
    pdf.set_font("helvetica", "B", 14)
    f_date = format_date_locale(from_date) if from_date else ""
    t_date = format_date_locale(to_date) if to_date else ""
    
    date_range_str = ""
    if from_date and to_date and from_date != to_date:
        date_range_str = f"{f_date} TO {t_date}"
    else:
        date_range_str = f_date or t_date or ""
        
    title = f"{tr('Attendance Report')} {date_range_str}".strip().upper()
    pdf.cell(0, 10, safe_pdf_str(title), align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("helvetica", "I", 9)
    pdf.cell(0, 5, safe_pdf_str(f"{tr('Generated:')} {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"), align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(8)
    
    # Precise Table Headers (Total width = 277mm)
    headers = [
        (tr("DATE"), 20), (tr("DEPARTMENT"), 27), (tr("MAT."), 10), (tr("LAST & FIRST NAME"), 45), 
        (tr("STATUS"), 8), (tr("CHECK IN 1"), 15), (tr("CHECK OUT 1"), 15), (tr("CHECK IN 2"), 15), 
        (tr("CHECK OUT 2"), 15), (tr("ATTENDANCE TIME"), 22), (tr("WORK TIME"), 20), (tr("DIFFERENCE"), 18), (tr("NOTE"), 47)
    ]
    
    # Draw Headers
    pdf.set_font("helvetica", "B", 7)
    pdf.set_fill_color(31, 78, 121)
    pdf.set_text_color(255, 255, 255)
    for h, w in headers:
        pdf.cell(w, 8, safe_pdf_str(h), border=1, align="C", fill=True)
    pdf.ln()
    
    pdf.set_text_color(0, 0, 0)
    
    # Grand tracking
    grand_total_jours = 0
    grand_status = {}
    grand_total_att = timedelta()
    grand_total_work = timedelta()

    # Grouping by Employee within Department
    for (dept, emp, reg), group in df.groupby(['department', 'employee', 'reg_number'], sort=False):
        pdf.set_font("helvetica", "", 7)
        fill = False
        
        # Row data accumulation for subtotals
        sub_p = 0
        emp_status = {}
        total_att = timedelta(); total_work = timedelta()
        
        for _, row in group.iterrows():
            try:
                pdf.set_fill_color(245, 245, 245) if fill else pdf.set_fill_color(255, 255, 255)
                
                # Extract and sanitize
                d = safe_pdf_str(row.get('date', ''))
                dp = safe_pdf_str(row.get('department', ''))
                mat = safe_pdf_str(row.get('reg_number', ''))
                nm = safe_pdf_str(row.get('employee', ''))
                st = safe_pdf_str(str(row.get('status', '')).strip())
                in1 = safe_pdf_str(row.get('time_in_1') or row.get('check_in') or '-')
                out1 = safe_pdf_str(row.get('time_out_1') or row.get('check_out') or '-')
                in2 = safe_pdf_str(row.get('time_in_2') or row.get('check_in_2') or '-')
                out2 = safe_pdf_str(row.get('time_out_2') or row.get('check_out_2') or '-')
                att = safe_pdf_str(row.get('attendance_time') or '-')
                wrk = safe_pdf_str(row.get('work_time') or '-')
                dif = safe_pdf_str(row.get('difference') or '-')
                nt = safe_pdf_str(row.get('note') or '')
                
                # Stats tracking
                if st and st != 'nan' and st != '-':
                    grand_status[st] = grand_status.get(st, 0) + 1
                    if st == 'P': sub_p += 1
                    else: emp_status[st] = emp_status.get(st, 0) + 1
                
                if att != '-': 
                    try: 
                        total_att += time_to_timedelta(att)
                        grand_total_att += time_to_timedelta(att)
                    except: pass
                if wrk != '-': 
                    try:
                        total_work += time_to_timedelta(wrk)
                        grand_total_work += time_to_timedelta(wrk)
                    except: pass
            except Exception as e:
                logger.error(f"Error processing detailed attendance row for PDF: {e}")
                continue

            # Draw cells with truncation
            pdf.cell(headers[0][1], 6, d, border=1, align="C", fill=True)
            pdf.cell(headers[1][1], 6, dp[:16] + ".." if len(dp) > 18 else dp, border=1, align="L", fill=True)
            pdf.cell(headers[2][1], 6, mat, border=1, align="C", fill=True)
            pdf.cell(headers[3][1], 6, nm[:23] + ".." if len(nm) > 25 else nm, border=1, align="L", fill=True)
            pdf.cell(headers[4][1], 6, st, border=1, align="C", fill=True)
            pdf.cell(headers[5][1], 6, in1, border=1, align="C", fill=True)
            pdf.cell(headers[6][1], 6, out1, border=1, align="C", fill=True)
            pdf.cell(headers[7][1], 6, in2, border=1, align="C", fill=True)
            pdf.cell(headers[8][1], 6, out2, border=1, align="C", fill=True)
            pdf.cell(headers[9][1], 6, att, border=1, align="C", fill=True)
            pdf.cell(headers[10][1], 6, wrk, border=1, align="C", fill=True)
            pdf.cell(headers[11][1], 6, dif, border=1, align="C", fill=True)
            pdf.cell(headers[12][1], 6, nt[:28] + ".." if len(nt) > 30 else nt, border=1, align="L", fill=True)
            pdf.ln()
            fill = not fill

        # Subtotal Row (Excel Style)
        grand_total_jours += sub_p
        
        pdf.set_font("helvetica", "B", 7)
        pdf.set_fill_color(225, 230, 245) # Light blue header-like fill
        
        # Row 1: Jours Présents + Duration Totals
        pdf.cell(headers[0][1]+headers[1][1]+headers[2][1], 6, "", border='LTB', fill=True)
        pdf.cell(headers[3][1], 6, safe_pdf_str(tr("Jours Présents")), border='TB', align="R", fill=True)
        pdf.cell(headers[4][1], 6, str(sub_p), border='TRB', align="C", fill=True)
        pdf.cell(headers[5][1]+headers[6][1]+headers[7][1]+headers[8][1], 6, "", border=1, fill=True)
        pdf.cell(headers[9][1], 6, format_timedelta(total_att), border=1, align="C", fill=True)
        pdf.cell(headers[10][1], 6, format_timedelta(total_work), border=1, align="C", fill=True)
        pdf.cell(headers[11][1]+headers[12][1], 6, "", border=1, fill=True)
        pdf.ln()

        # Supplementary status rows (dynamic)
        for st_name, count in emp_status.items():
            if count > 0 and st_name != '-' and st_name != 'nan':
                pdf.cell(headers[0][1]+headers[1][1]+headers[2][1], 6, "", border=0)
                pdf.cell(headers[3][1], 6, safe_pdf_str(f"{tr('Total')} {st_name}"), border='LTB', align="R", fill=False)
                pdf.cell(headers[4][1], 6, str(count), border='TRB', align="C", fill=False)
                pdf.ln()
        
        pdf.ln(3) # Group separator
        
    # GRAND TOTAL
    pdf.set_font("helvetica", "B", 7)
    pdf.set_fill_color(210, 215, 225)
    
    pdf.cell(headers[0][1]+headers[1][1]+headers[2][1], 6, "", border='LTB', fill=True)
    pdf.cell(headers[3][1], 6, safe_pdf_str(tr("Total Jours Présents")), border='TB', align="R", fill=True)
    pdf.cell(headers[4][1], 6, str(grand_total_jours), border='TRB', align="C", fill=True)
    pdf.cell(headers[5][1]+headers[6][1]+headers[7][1]+headers[8][1], 6, "", border=1, fill=True)
    pdf.cell(headers[9][1], 6, format_timedelta(grand_total_att), border=1, align="C", fill=True)
    pdf.cell(headers[10][1], 6, format_timedelta(grand_total_work), border=1, align="C", fill=True)
    pdf.cell(headers[11][1]+headers[12][1], 6, "", border=1, fill=True)
    pdf.ln()

    for st_name, count in grand_status.items():
        if count > 0 and st_name != 'P' and st_name != '-' and st_name != 'nan':
            pdf.cell(headers[0][1]+headers[1][1]+headers[2][1], 6, "", border=0)
            pdf.cell(headers[3][1], 6, safe_pdf_str(f"{tr('Total')} {st_name}"), border='LTB', align="R", fill=True)
            pdf.cell(headers[4][1], 6, str(count), border='TRB', align="C", fill=True)
            pdf.ln()

    try:
        pdf.output(filepath)
    except Exception as e:
        raise RuntimeError(f"Could not save PDF file. Is it already open? Error: {e}")


def generate_daily_detailed_attendance_pdf(data, filepath, from_date=None, to_date=None):
    if not data:
        raise ValueError("No data available to export.")
    
    df = pd.DataFrame(data)
    
    sort_key = 'raw_date' if 'raw_date' in df.columns else 'date'
    df.sort_values(by=[sort_key, 'department', 'employee'], inplace=True)
    if 'id' in df.columns:
        mask_real = df['id'].notna() & (df['id'] != -1)
        df_real = df[mask_real].drop_duplicates(subset=['id'])
        df_synth = df[~mask_real]
        df = pd.concat([df_real, df_synth], ignore_index=True)
        df.sort_values(by=[sort_key, 'department', 'employee'], inplace=True)
    elif not df.empty:
        df.drop_duplicates(inplace=True)

    grouped_date = df.groupby(sort_key, sort=False)

    class PDF(FPDF):
        def footer(self):
            self.set_y(-15)
            self.set_font("helvetica", "I", 8)
            self.cell(0, 10, f"Page {self.page_no()}", align="C")

    pdf = PDF(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    
    # Header logic
    logo_path = _get_company_logo_path()
    if logo_path:
        try: pdf.image(logo_path, x=10, y=8, h=20)
        except: pass
            
    pdf.set_xy(10, 10)
    pdf.set_font("helvetica", "B", 14)
    f_date = format_date_locale(from_date) if from_date else ""
    t_date = format_date_locale(to_date) if to_date else ""
    
    date_range_str = ""
    if from_date and to_date and from_date != to_date:
        date_range_str = f"{f_date} TO {t_date}"
    else:
        date_range_str = f_date or t_date or ""
        
    title = f"{tr('Daily Attendance Report')} {date_range_str}".strip().upper()
    pdf.cell(0, 10, safe_pdf_str(title), align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("helvetica", "I", 9)
    pdf.cell(0, 5, safe_pdf_str(f"{tr('Generated:')} {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"), align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(8)
    
    # Precise Table Headers (Total width = 277mm)
    headers = [
        (tr("DATE"), 20), (tr("DEPARTMENT"), 27), (tr("MAT."), 10), (tr("LAST & FIRST NAME"), 45), 
        (tr("STATUS"), 8), (tr("CHECK IN 1"), 15), (tr("CHECK OUT 1"), 15), (tr("CHECK IN 2"), 15), 
        (tr("CHECK OUT 2"), 15), (tr("ATTENDANCE TIME"), 22), (tr("WORK TIME"), 20), (tr("DIFFERENCE"), 18), (tr("NOTE"), 47)
    ]
    
    # Draw Headers
    pdf.set_font("helvetica", "B", 7)
    pdf.set_fill_color(31, 78, 121)
    pdf.set_text_color(255, 255, 255)
    for h, w in headers:
        pdf.cell(w, 8, safe_pdf_str(h), border=1, align="C", fill=True)
    pdf.ln()
    
    pdf.set_text_color(0, 0, 0)

    grand_total_jours = 0
    grand_total_travail = timedelta()
    grand_total_presence = timedelta()
    
    for date_key, date_group in grouped_date:
        date_travail = timedelta()
        date_presence = timedelta()
        date_jours_presents = 0
        date_status = {}
        
        grouped_dept = date_group.groupby('department', sort=False)
        for dept_key, dept_group in grouped_dept:
            pdf.set_font("helvetica", "", 7)
            fill = False
            
            dept_travail = timedelta()
            dept_presence = timedelta()
            dept_jours_presents = 0
            dept_status = {}
            
            for _, row in dept_group.iterrows():
                pdf.set_fill_color(245, 245, 245) if fill else pdf.set_fill_color(255, 255, 255)
                
                # Extract and sanitize
                d = safe_pdf_str(format_date_locale(row.get('date', ''), short=True))
                dp = safe_pdf_str(str(row.get('department', '')))
                mat = safe_pdf_str(str(row.get('reg_number', '')))
                nm = safe_pdf_str(str(row.get('employee', '')))
                st = safe_pdf_str(str(row.get('status', '')).strip())
                in1 = safe_pdf_str(str(row.get('time_in_1') or str(row.get('check_in', '-')))[:5] if row.get('check_in', '-') != '-' else '-')
                out1 = safe_pdf_str(str(row.get('time_out_1') or str(row.get('check_out', '-')))[:5] if row.get('check_out', '-') != '-' else '-')
                in2 = safe_pdf_str(str(row.get('time_in_2') or str(row.get('check_in_2', '-')))[:5] if row.get('check_in_2', '-') != '-' else '-')
                out2 = safe_pdf_str(str(row.get('time_out_2') or str(row.get('check_out_2', '-')))[:5] if row.get('check_out_2', '-') != '-' else '-')
                att = safe_pdf_str(str(row.get('attendance_time') or '-'))
                wrk = safe_pdf_str(str(row.get('work_time') or '-'))
                dif = safe_pdf_str(str(row.get('difference') or '-'))
                nt = safe_pdf_str(str(row.get('note') or ''))
                
                if st and st != 'nan' and st != '-':
                    dept_status[st] = dept_status.get(st, 0) + 1
                    date_status[st] = date_status.get(st, 0) + 1
                    
                if att != '-': 
                    dept_jours_presents += 1
                    try:
                        dept_presence += time_to_timedelta(att)
                        date_presence += time_to_timedelta(att)
                        grand_total_presence += time_to_timedelta(att)
                    except: pass
                if wrk != '-': 
                    try:
                        dept_travail += time_to_timedelta(wrk)
                        date_travail += time_to_timedelta(wrk)
                        grand_total_travail += time_to_timedelta(wrk)
                    except: pass

                # Draw cells with truncation
                pdf.cell(headers[0][1], 6, d, border=1, align="C", fill=True)
                pdf.cell(headers[1][1], 6, dp[:16] + ".." if len(dp) > 18 else dp, border=1, align="L", fill=True)
                pdf.cell(headers[2][1], 6, mat, border=1, align="C", fill=True)
                pdf.cell(headers[3][1], 6, nm[:23] + ".." if len(nm) > 25 else nm, border=1, align="L", fill=True)
                pdf.cell(headers[4][1], 6, st, border=1, align="C", fill=True)
                pdf.cell(headers[5][1], 6, in1, border=1, align="C", fill=True)
                pdf.cell(headers[6][1], 6, out1, border=1, align="C", fill=True)
                pdf.cell(headers[7][1], 6, in2, border=1, align="C", fill=True)
                pdf.cell(headers[8][1], 6, out2, border=1, align="C", fill=True)
                pdf.cell(headers[9][1], 6, att, border=1, align="C", fill=True)
                pdf.cell(headers[10][1], 6, wrk, border=1, align="C", fill=True)
                pdf.cell(headers[11][1], 6, dif, border=1, align="C", fill=True)
                pdf.cell(headers[12][1], 6, nt[:28] + ".." if len(nt) > 30 else nt, border=1, align="L", fill=True)
                pdf.ln()
                fill = not fill

            # SUB-TOTAL for Department
            date_jours_presents += dept_jours_presents
            grand_total_jours += dept_jours_presents
            
            pdf.set_font("helvetica", "B", 7)
            pdf.set_fill_color(225, 230, 245) # Light blue header-like fill
            
            # Subtotal Row Dept
            pdf.cell(headers[0][1]+headers[1][1]+headers[2][1], 6, "", border='LTB', fill=True)
            pdf.cell(headers[3][1], 6, safe_pdf_str(f"{tr('Total')} {dept_key}"), border='TB', align="R", fill=True)
            pdf.cell(headers[4][1], 6, str(dept_jours_presents), border='TRB', align="C", fill=True)
            pdf.cell(headers[5][1]+headers[6][1]+headers[7][1]+headers[8][1], 6, "", border=1, fill=True)
            pdf.cell(headers[9][1], 6, format_timedelta(dept_presence), border=1, align="C", fill=True)
            pdf.cell(headers[10][1], 6, format_timedelta(dept_travail), border=1, align="C", fill=True)
            pdf.cell(headers[11][1]+headers[12][1], 6, "", border=1, fill=True)
            pdf.ln()
            
            for st, count in dept_status.items():
                if st == 'P' or count == 0: continue
                pdf.cell(headers[0][1]+headers[1][1]+headers[2][1], 6, "", border=0)
                pdf.cell(headers[3][1], 6, safe_pdf_str(f"{tr('Total')} {st}"), border='LTB', align="R", fill=False)
                pdf.cell(headers[4][1], 6, str(count), border='TRB', align="C", fill=False)
                pdf.ln()
            
            pdf.ln(2)

        # DATE LEVEL TOTAL
        group_date_str = safe_pdf_str(format_date_locale(date_key))
        pdf.set_font("helvetica", "B", 7)
        pdf.set_fill_color(210, 215, 225)
        
        pdf.cell(headers[0][1]+headers[1][1]+headers[2][1], 6, "", border='LTB', fill=True)
        pdf.cell(headers[3][1], 6, safe_pdf_str(f"{tr('Total pour')} {group_date_str}"), border='TB', align="R", fill=True)
        pdf.cell(headers[4][1], 6, str(date_jours_presents), border='TRB', align="C", fill=True)
        pdf.cell(headers[5][1]+headers[6][1]+headers[7][1]+headers[8][1], 6, "", border=1, fill=True)
        pdf.cell(headers[9][1], 6, format_timedelta(date_presence), border=1, align="C", fill=True)
        pdf.cell(headers[10][1], 6, format_timedelta(date_travail), border=1, align="C", fill=True)
        pdf.cell(headers[11][1]+headers[12][1], 6, "", border=1, fill=True)
        pdf.ln()
        pdf.ln(5)

    try:
        pdf.output(filepath)
    except Exception as e:
        raise RuntimeError(f"Could not save PDF file. Is it already open? Error: {e}")
